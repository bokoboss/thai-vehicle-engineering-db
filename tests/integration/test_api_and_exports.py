from __future__ import annotations

import csv
import io

from openpyxl import load_workbook


def test_health_and_vehicle_search(client):
    assert client.get("/healthz").json() == {"status": "ok"}
    response = client.get("/api/vehicles", params={"q": "nominal-width"})
    assert response.status_code == 200
    assert response.json()["count"] == 1
    assert response.json()["items"][0]["stable_vehicle_code"] == "FIXTURE-AVT-TRACK-SCREENING"
    filtered = client.get("/api/vehicles", params={"manufacturer": "Phase 0 Contract Fixtures"})
    assert filtered.status_code == 200
    assert filtered.json()["count"] == 16


def test_root_redirects_to_catalog_start_page(client):
    response = client.get("/", follow_redirects=False)

    assert response.status_code == 307
    assert response.headers["location"] == "/vehicles"


def test_detail_api_and_html_expose_state_and_evidence(client):
    response = client.get("/api/vehicles/FIXTURE-WIDTH-UNSPECIFIED")
    assert response.status_code == 200
    body = response.json()
    width = next(item for item in body["values"] if item["parameter_code"] == "overall_width_reported_mm")
    assert width["availability_state"] == "AVAILABLE"
    assert width["semantic_metadata"]["width_envelope_definition"] == "OEM_UNSPECIFIED"
    assert body["assessments"][0]["parameter_code"] == "overall_width_body_mm"
    html = client.get("/vehicles/FIXTURE-WIDTH-UNSPECIFIED")
    assert html.status_code == 200
    assert "Unknown / research assessments" in html.text
    assert "Overall Width" in html.text


def test_detail_exposes_load_scope_and_source_document_identity(client):
    response = client.get("/api/vehicles/FIXTURE-STATIC-LOADED-RADIUS")
    assert response.status_code == 200
    body = response.json()
    radius = next(
        item for item in body["values"] if item["parameter_code"] == "static_loaded_tyre_radius_front_mm"
    )
    assert radius["load_condition"]["name"] == "Measured static-loaded radius condition"
    assert radius["load_condition"]["mass_basis"] == "KERB"
    assert radius["observations"][0]["source_title"] == "Static-loaded tyre radius fixture"
    assert radius["observations"][0]["source_publisher"] == "Phase 0 deterministic fixture author"
    assert radius["observations"][0]["source_url"] == (
        "https://example.invalid/thai-vehicle-engineering-db/fixtures/fixture-static-loaded-radius"
    )
    html = client.get("/vehicles/FIXTURE-STATIC-LOADED-RADIUS")
    assert html.status_code == 200
    assert "Measured static-loaded radius condition" in html.text
    assert "Static-loaded tyre radius fixture" in html.text
    assert "https://example.invalid/thai-vehicle-engineering-db/fixtures/fixture-static-loaded-radius" in html.text


def test_issues_and_compare_pages_are_available(client):
    issues = client.get("/api/issues")
    assert issues.status_code == 200
    assert any(item["code"] == "AVT_READY" for item in issues.json()["items"])
    compare = client.get("/compare", params={"codes": "FIXTURE-PRIMARY-PUBLISHED,FIXTURE-AVT-TRACK-DIRECT"})
    assert compare.status_code == 200
    assert "Direct AVT outer-face track" in compare.text


def test_csv_export_preserves_states_sources_and_assessments(client):
    response = client.get("/exports/vehicles.csv", params={"codes": "FIXTURE-CONFLICTING-VALUE,FIXTURE-UNKNOWN-ASSESSMENT"})
    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8-sig"))))
    conflict = [row for row in rows if row["parameter_code"] == "overall_length_mm"]
    assert len(conflict) == 2
    assert all(row["resolution_state"] == "CONFLICTING" for row in conflict)
    unknown = next(row for row in rows if row["value_kind"] == "assessment")
    assert unknown["availability_state"] == "NOT_FOUND_AFTER_SEARCH"
    assert unknown["normalized_value"] == ""
    assert unknown["assessment_reason"]
    assert all(row["source_observation_ids"] for row in conflict)


def test_xlsx_export_is_readable_and_has_evidence_aware_headers(client):
    response = client.get("/exports/vehicles.xlsx", params={"codes": "FIXTURE-AVT-TRACK-DIRECT"})
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    sheet = workbook["Engineering Data"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert "source_observation_ids" in headers
    assert "resolution_state" in headers
    assert "derivation_rule_version" in headers
    stable_code_index = headers.index("stable_vehicle_code")
    assert any(row[stable_code_index].value == "FIXTURE-AVT-TRACK-DIRECT" for row in sheet.iter_rows(min_row=2))
