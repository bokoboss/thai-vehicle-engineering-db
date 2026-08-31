from __future__ import annotations

import csv
import json
from copy import deepcopy
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, sessionmaker

from app.curate.__main__ import main as curate_main
from app.curate.loader import import_manifest, initialize_registry
from app.curate.validation import CurationError, load_manifest, source_notes, validate_manifest
from app.db.base import Base
from app.db.models import (
    ConflictDecision,
    EvidenceLink,
    Manufacturer,
    NormalizedValue,
    ParameterAssessment,
    ParameterDefinition,
    ReadinessResult,
    SourceDocument,
    SourceObservation,
    VehicleConfiguration,
    VehicleFitment,
    VehicleModel,
)
from app.db.session import make_engine
from app.domain.enums import ReadinessType
from app.domain.schemas import SourceDocumentCreate
from app.exports.exporter import EXPORT_COLUMNS, csv_bytes, export_rows, xlsx_bytes
from app.seed.registry import seed_registry
from app.services.foundation import create_source_document


MANIFEST_DIR = Path(__file__).resolve().parents[2] / "data" / "curation" / "manifests" / "sentinel"
BYD_MANIFEST = MANIFEST_DIR / "byd_atto3_my24_extended_local_v1.json"
TRITON_MANIFEST = MANIFEST_DIR / "mitsubishi_triton_ultra_4wd_at_2023_v1.json"
VOLVO_MANIFEST = MANIFEST_DIR / "volvo_ex30_ultra_smer_my2026_v1.json"


@pytest.fixture
def curated_session(tmp_path: Path):
    database_path = tmp_path / "curated.sqlite"
    engine = make_engine(f"sqlite:///{database_path}")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    session = factory()
    initialize_registry(session)
    try:
        yield session
    finally:
        session.rollback()
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def manifest_dict(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def count(session: Session, model: type) -> int:
    return int(session.scalar(select(func.count()).select_from(model)) or 0)


def persisted_counts(session: Session) -> dict[str, int]:
    return {
        "vehicles": count(session, VehicleConfiguration),
        "manufacturers": count(session, Manufacturer),
        "models": count(session, VehicleModel),
        "sources": count(session, SourceDocument),
        "observations": count(session, SourceObservation),
        "values": count(session, NormalizedValue),
        "assessments": count(session, ParameterAssessment),
        "fitments": count(session, VehicleFitment),
        "conflicts": count(session, ConflictDecision),
    }


def vehicle(session: Session, stable_code: str) -> VehicleConfiguration:
    result = session.scalar(
        select(VehicleConfiguration).where(VehicleConfiguration.stable_vehicle_code == stable_code)
    )
    assert result is not None
    return result


def values_for(session: Session, config: VehicleConfiguration, parameter_code: str) -> list[NormalizedValue]:
    return list(
        session.scalars(
            select(NormalizedValue)
            .join(NormalizedValue.parameter_definition)
            .where(
                NormalizedValue.vehicle_configuration_id == config.id,
                ParameterDefinition.parameter_code == parameter_code,
            )
            .order_by(NormalizedValue.id)
        ).all()
    )


def readiness_for(
    session: Session,
    config: VehicleConfiguration,
    readiness_type: ReadinessType,
    *,
    fitment: VehicleFitment | None = None,
) -> ReadinessResult:
    statement = select(ReadinessResult).where(
        ReadinessResult.vehicle_configuration_id == config.id,
        ReadinessResult.readiness_type == readiness_type.value,
    )
    if fitment is None:
        statement = statement.where(ReadinessResult.vehicle_fitment_id.is_(None))
    else:
        statement = statement.where(ReadinessResult.vehicle_fitment_id == fitment.id)
    result = session.scalar(statement)
    assert result is not None
    return result


def test_init_seeds_registry_only_and_no_phase0_fixtures(curated_session: Session):
    assert count(curated_session, ParameterDefinition) == 48
    assert count(curated_session, VehicleConfiguration) == 0
    assert count(curated_session, SourceDocument) == 0

    # Re-running init is safe and still does not create a vehicle fixture.
    assert initialize_registry(curated_session) == 48
    assert count(curated_session, VehicleConfiguration) == 0


@pytest.mark.parametrize(
    "name,mutate",
    [
        ("unknown parameter", lambda manifest: manifest["values"][0].update(parameter_code="not_in_registry")),
        ("undeclared source", lambda manifest: manifest["observations"][0].update(source_code="MISSING_SOURCE")),
        ("missing evidence", lambda manifest: manifest["values"][0].update(evidence_observation_codes=[])),
        (
            "missing load condition",
            lambda manifest: manifest["values"][6].update(load_condition_code="MISSING_LOAD"),
        ),
        (
            "missing fitment",
            lambda manifest: manifest["values"][0].update(fitment_code="MISSING_FITMENT"),
        ),
        (
            "direct derived",
            lambda manifest: manifest["values"][0].update(evidence_method="DERIVED"),
        ),
        (
            "direct estimated",
            lambda manifest: manifest["values"][0].update(evidence_method="ESTIMATED"),
        ),
        (
            "wrong registry primitive",
            lambda manifest: manifest["values"][0].update(value="4455"),
        ),
        (
            "invalid identity time basis",
            lambda manifest: manifest["vehicle"].update(identity_time_basis="UNKNOWN"),
        ),
    ],
)
def test_manifest_negative_gates_fail_before_a_successful_import(
    curated_session: Session,
    name: str,
    mutate,
):
    del name
    manifest = manifest_dict(BYD_MANIFEST)
    mutate(manifest)
    with pytest.raises(CurationError):
        import_manifest(curated_session, manifest)
    assert persisted_counts(curated_session) == {
        "vehicles": 0,
        "manufacturers": 0,
        "models": 0,
        "sources": 0,
        "observations": 0,
        "values": 0,
        "assessments": 0,
        "fitments": 0,
        "conflicts": 0,
    }


def test_strict_manifest_rejects_unknown_top_level_fields(curated_session: Session):
    manifest = manifest_dict(BYD_MANIFEST)
    manifest["unexpected"] = True
    with pytest.raises(CurationError, match="unexpected"):
        import_manifest(curated_session, manifest)


def test_duplicate_stable_vehicle_code_is_create_only(curated_session: Session):
    manifest = manifest_dict(BYD_MANIFEST)
    import_manifest(curated_session, manifest)
    before = persisted_counts(curated_session)

    with pytest.raises(CurationError, match="CREATE_ONLY"):
        import_manifest(curated_session, deepcopy(manifest))
    assert persisted_counts(curated_session) == before


def _source_create_payload(source) -> SourceDocumentCreate:
    return SourceDocumentCreate(
        source_code=source.source_code,
        title=source.title,
        publisher=source.publisher,
        authority_class=source.authority_class.value,
        source_type=source.source_type.value,
        market_code=source.market_code,
        publication_year=source.publication_year,
        model_year_from=source.model_year_from,
        model_year_to=source.model_year_to,
        url=source.url,
        retrieved_at=source.retrieved_at,
        local_snapshot_reference=source.local_snapshot_reference,
        content_hash=source.content_hash,
        page_section_default=source.page_section_default,
        access_licensing_notes=source.access_licensing_notes,
        applicability_notes=source.applicability_notes,
        archival_status=source.archival_status,
        notes=source_notes(source),
    )


def test_compatible_source_is_reused_without_mutation(curated_session: Session):
    manifest = load_manifest(BYD_MANIFEST)
    stored_source = _source_create_payload(manifest.sources[0])
    stored_source.notes = f"{stored_source.notes}\nUnrelated existing note"
    create_source_document(curated_session, stored_source)
    curated_session.commit()

    report = import_manifest(curated_session, manifest)
    assert report.sources_created == 1
    assert report.sources_reused == 1
    assert count(curated_session, SourceDocument) == 2
    source = curated_session.scalar(
        select(SourceDocument).where(SourceDocument.source_code == manifest.sources[0].source_code)
    )
    assert source is not None
    assert source.notes is not None
    assert "source_subtype_raw: WEB_SPECIFICATION" in source.notes


def test_prefix_compatible_source_subtype_mismatch_fails_without_vehicle_data(
    curated_session: Session,
):
    manifest = load_manifest(BYD_MANIFEST)
    stored_source = _source_create_payload(manifest.sources[0])
    stored_source.notes = "source_subtype_raw: WEB_SPECIFICATION_SCREENSHOT\nUnrelated existing note"
    create_source_document(curated_session, stored_source)
    curated_session.commit()
    stored_notes = stored_source.notes

    bad = manifest_dict(BYD_MANIFEST)
    bad["vehicle"]["stable_vehicle_code"] = "th-byd-source-subtype-conflict-test"
    with pytest.raises(CurationError, match="source_subtype_raw"):
        import_manifest(curated_session, bad)

    source = curated_session.scalar(
        select(SourceDocument).where(SourceDocument.source_code == manifest.sources[0].source_code)
    )
    assert source is not None
    assert source.notes == stored_notes
    assert count(curated_session, VehicleConfiguration) == 0
    assert count(curated_session, SourceObservation) == 0
    assert count(curated_session, NormalizedValue) == 0


def test_incompatible_existing_source_fails_without_mutating_it(curated_session: Session):
    manifest = load_manifest(BYD_MANIFEST)
    create_source_document(curated_session, _source_create_payload(manifest.sources[0]))
    curated_session.commit()

    bad = manifest_dict(BYD_MANIFEST)
    bad["vehicle"]["stable_vehicle_code"] = "th-byd-source-conflict-test"
    bad["sources"][0]["publisher"] = "Untrusted publisher"
    with pytest.raises(CurationError, match="conflicts with existing source metadata"):
        import_manifest(curated_session, bad)

    source = curated_session.scalar(
        select(SourceDocument).where(SourceDocument.source_code == manifest.sources[0].source_code)
    )
    assert source is not None
    assert source.publisher == "REVER Automotive"
    assert count(curated_session, VehicleConfiguration) == 0


def test_incompatible_existing_manufacturer_or_model_fails_without_mutation(curated_session: Session):
    manifest = manifest_dict(BYD_MANIFEST)
    manufacturer = Manufacturer(canonical_name="BYD", display_name="BYD Legacy")
    curated_session.add(manufacturer)
    curated_session.commit()

    with pytest.raises(CurationError, match="manufacturer BYD"):
        import_manifest(curated_session, manifest)
    assert curated_session.scalar(select(Manufacturer).where(Manufacturer.canonical_name == "BYD")).display_name == "BYD Legacy"
    assert count(curated_session, VehicleConfiguration) == 0

    curated_session.query(Manufacturer).delete()
    curated_session.commit()
    manufacturer = Manufacturer(canonical_name="BYD", display_name="BYD")
    curated_session.add(manufacturer)
    curated_session.flush()
    curated_session.add(
        VehicleModel(
            manufacturer_id=manufacturer.id,
            canonical_model_name="atto3",
            display_model_name="ATTO 3 Legacy",
        )
    )
    curated_session.commit()

    with pytest.raises(CurationError, match="model atto3"):
        import_manifest(curated_session, manifest)
    assert curated_session.scalar(
        select(VehicleModel).where(VehicleModel.canonical_model_name == "atto3")
    ).display_model_name == "ATTO 3 Legacy"
    assert count(curated_session, VehicleConfiguration) == 0


def test_final_service_failure_rolls_back_the_whole_manifest(curated_session: Session):
    manifest = manifest_dict(BYD_MANIFEST)
    final_value = manifest["values"][-1]
    final_value.update(
        parameter_code="clearance_value_mm",
        value=18,
        canonical_unit="mm",
        semantic_metadata={"clearance_type": "NOT_A_CONTROLLED_CLEARANCE"},
        load_condition_code="UNLADEN_OEM",
    )

    with pytest.raises(CurationError, match="clearance"):
        import_manifest(curated_session, manifest)

    assert persisted_counts(curated_session) == {
        "vehicles": 0,
        "manufacturers": 0,
        "models": 0,
        "sources": 0,
        "observations": 0,
        "values": 0,
        "assessments": 0,
        "fitments": 0,
        "conflicts": 0,
    }


def test_validate_and_dry_run_do_not_persist(curated_session: Session):
    manifest = load_manifest(BYD_MANIFEST)
    before = persisted_counts(curated_session)

    validate_manifest(curated_session, manifest)
    assert persisted_counts(curated_session) == before
    report = import_manifest(curated_session, manifest, dry_run=True)
    assert report.status == "DRY_RUN"
    assert report.normalized_values == 15
    assert persisted_counts(curated_session) == before


@pytest.mark.parametrize("manifest_path", [BYD_MANIFEST, TRITON_MANIFEST, VOLVO_MANIFEST])
def test_cli_validate_runs_full_dry_run_path_for_all_sentinels(
    curated_session: Session,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    manifest_path: Path,
):
    monkeypatch.setattr("app.curate.__main__.SessionLocal", lambda: curated_session)
    before = persisted_counts(curated_session)

    assert curate_main(["validate", str(manifest_path)]) == 0
    report = json.loads(capsys.readouterr().out)
    assert report["status"] == "PASS"
    assert persisted_counts(curated_session) == before


def test_cli_validate_rejects_invalid_controlled_clearance_without_persistence(
    curated_session: Session,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    tmp_path: Path,
):
    manifest = manifest_dict(BYD_MANIFEST)
    manifest["values"][-1].update(
        parameter_code="clearance_value_mm",
        value=18,
        canonical_unit="mm",
        semantic_metadata={"clearance_type": "NOT_A_CONTROLLED_CLEARANCE"},
        load_condition_code="UNLADEN_OEM",
    )
    manifest_path = tmp_path / "invalid-clearance.json"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    monkeypatch.setattr("app.curate.__main__.SessionLocal", lambda: curated_session)
    before = persisted_counts(curated_session)

    assert curate_main(["validate", str(manifest_path)]) == 1
    assert "clearance" in capsys.readouterr().out.lower()
    assert persisted_counts(curated_session) == before


def test_sentinel_imports_preserve_lineage_readiness_and_exports(curated_session: Session):
    reports = [
        import_manifest(curated_session, load_manifest(path))
        for path in (BYD_MANIFEST, TRITON_MANIFEST, VOLVO_MANIFEST)
    ]
    assert [report.normalized_values for report in reports] == [15, 12, 15]
    assert [report.observations for report in reports] == [13, 7, 14]
    assert [report.assessments for report in reports] == [2, 2, 3]
    assert [report.loads for report in reports] == [2, 1, 1]
    assert [report.fitments for report in reports] == [0, 0, 1]
    assert persisted_counts(curated_session)["vehicles"] == 3

    byd = vehicle(curated_session, "th-byd-atto3-my24-extended-local")
    assert byd.model_year_from == 2024
    assert byd.identity_time_basis == "MODEL_YEAR"
    byd_clearance = values_for(curated_session, byd, "clearance_value_mm")
    assert sorted(value.numeric_value for value in byd_clearance) == [150.0, 175.0]
    assert len({value.load_condition_id for value in byd_clearance}) == 2
    assert {value.load_condition.mass_basis for value in byd_clearance} == {"UNLADEN", "OEM_LADEN"}
    assert {value.evidence_links[0].source_observation.raw_label for value in byd_clearance} == {
        "Ground Clearance Unladen",
        "Ground Clearance Laden",
    }
    assert {value.evidence_links[0].source_observation.source_document.source_code for value in byd_clearance} == {
        "REVER_TH_ATTO3_SPEC_CURRENT"
    }
    assert {assessment.parameter_definition.parameter_code for assessment in byd.parameter_assessments} == {
        "avt_front_outer_face_track_mm",
        "avt_rear_outer_face_track_mm",
    }
    assert readiness_for(curated_session, byd, ReadinessType.IDENTITY_RESOLVED).status == "READY"
    assert readiness_for(curated_session, byd, ReadinessType.AVT_READY).status == "NOT_READY"

    triton = vehicle(curated_session, "th-mitsubishi-triton-ultra-4wd-at-release-2023")
    assert triton.model_year_from is None
    assert triton.model_year_to is None
    assert triton.identity_time_basis == "EDITION_RELEASE"
    assert triton.identity_time_label_raw is not None
    assert readiness_for(curated_session, triton, ReadinessType.IDENTITY_RESOLVED).status == "READY"
    triton_turning = values_for(curated_session, triton, "turning_radius_normalized_m")[0]
    assert triton_turning.numeric_value == 6.2
    assert triton_turning.semantic_metadata == {
        "turning_radius_or_diameter": "RADIUS",
        "turning_reference": "OEM_UNSPECIFIED",
        "turning_axle_scope": "OEM_UNSPECIFIED",
    }
    assert readiness_for(curated_session, triton, ReadinessType.AVT_READY).status == "NOT_READY"

    volvo = vehicle(curated_session, "th-volvo-ex30-ultra-smer-my2026-19")
    assert volvo.model_year_from == 2026
    fitment = volvo.fitments[0]
    assert fitment.fitment_code == "WHEEL19"
    for value in values_for(curated_session, volvo, "front_tyre_size_text") + values_for(
        curated_session, volvo, "rear_tyre_size_text"
    ):
        assert value.vehicle_fitment_id == fitment.id
    assert {
        values_for(curated_session, volvo, code)[0].numeric_value
        for code in (
            "overall_width_body_mm",
            "overall_width_including_mirrors_mm",
            "overall_width_mirrors_folded_mm",
        )
    } == {1838.0, 2032.0, 1940.0}
    assert values_for(curated_session, volvo, "overall_height_mm")[0].load_condition.name == "Kerb weight plus one person"
    assert values_for(curated_session, volvo, "clearance_value_mm")[0].numeric_value == 171.0
    turning_text = values_for(curated_session, volvo, "oem_turning_value_text")
    assert {value.text_value for value in turning_text} == {"10.7 m", "11 m"}
    assert {value.resolution_state for value in turning_text} == {"CONFLICTING"}
    assert values_for(curated_session, volvo, "turning_radius_normalized_m") == []
    assert [
        assessment.availability_state
        for assessment in volvo.parameter_assessments
        if assessment.parameter_definition.parameter_code == "turning_radius_normalized_m"
    ] == ["UNKNOWN"]
    assert count(curated_session, ConflictDecision) == 0
    assert readiness_for(curated_session, volvo, ReadinessType.IDENTITY_RESOLVED).status == "READY"
    assert readiness_for(curated_session, volvo, ReadinessType.AVT_READY).status == "NOT_READY"
    assert readiness_for(curated_session, volvo, ReadinessType.AVT_READY, fitment=fitment).status == "NOT_READY"

    # Representative source -> observation -> normalized value lineage for all sentinels.
    lineage_expectations = {
        byd: ("overall_length_mm", "Overall Length", "REVER_TH_ATTO3_SPEC_CURRENT"),
        triton: ("overall_length_mm", "Overall dimension with rear bumper", "MITSUBISHI_TH_TRITON_DOUBLECAB_BROCHURE_202311"),
        volvo: ("overall_width_body_mm", "Width", "VOLVO_TH_EX30_SUPPORT_DIMENSIONS"),
    }
    for config, (parameter_code, raw_label, source_code) in lineage_expectations.items():
        value = values_for(curated_session, config, parameter_code)[0]
        assert len(value.evidence_links) == 1
        link: EvidenceLink = value.evidence_links[0]
        assert link.source_observation.vehicle_configuration_id == config.id
        assert link.source_observation.raw_label == raw_label
        assert link.source_observation.source_document.source_code == source_code

    rows = export_rows(
        curated_session,
        [
            byd.stable_vehicle_code,
            triton.stable_vehicle_code,
            volvo.stable_vehicle_code,
        ],
    )
    assert rows
    assert set(EXPORT_COLUMNS).issubset(rows[0])
    csv_rows = list(csv.DictReader(csv_bytes(rows).decode("utf-8-sig").splitlines()))
    assert {row["stable_vehicle_code"] for row in csv_rows} == {
        byd.stable_vehicle_code,
        triton.stable_vehicle_code,
        volvo.stable_vehicle_code,
    }
    assert {row["identity_time_basis"] for row in csv_rows} == {"MODEL_YEAR", "EDITION_RELEASE"}
    assert any(
        row["stable_vehicle_code"] == volvo.stable_vehicle_code
        and row["resolution_state"] == "CONFLICTING"
        and row["source_document_codes"]
        for row in csv_rows
    )
    workbook = load_workbook(BytesIO(xlsx_bytes(rows)), read_only=True)
    worksheet = workbook["Engineering Data"]
    headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    assert headers == EXPORT_COLUMNS
    workbook.close()


def test_structured_axle_steering_geometry_and_conflict_decision_paths(curated_session: Session):
    manifest = manifest_dict(BYD_MANIFEST)
    manifest["axles"] = [
        {
            "axle_code": "FRONT",
            "axle_role": "FRONT",
            "axle_index": 0,
            "longitudinal_position_mm": 0,
            "driven": False,
            "steered": True,
        },
        {
            "axle_code": "REAR",
            "axle_role": "REAR",
            "axle_index": 1,
            "longitudinal_position_mm": 2720,
            "driven": True,
            "steered": False,
        },
    ]
    manifest["steering_relations"] = [
        {
            "axle_code": "FRONT",
            "steering_role": "PRIMARY",
            "linkage_type": "FIXED_RATIO",
            "phase_behavior": "SAME_PHASE",
            "source_observation_code": "OBS_TURNING",
        }
    ]
    manifest["geometry_assets"] = [
        {
            "geometry_code": "AXLE_DATUM",
            "geometry_role": "AXLE_DATUM_GEOMETRY",
            "representation_type": "PARAMETRIC",
            "geometry_data": {"wheelbase_mm": 2720},
            "unit": "mm",
            "coordinate_system_version": "vehicle-fixed-v1",
            "geometry_method": "OEM_DIMENSION_DRAWING",
            "geometry_fidelity": "MEDIUM",
            "source_code": "REVER_TH_ATTO3_SPEC_CURRENT",
            "uncertainty_description": "Dimension diagram is an axle datum callout, not an AVT body profile.",
        }
    ]
    conflict_value = deepcopy(manifest["values"][0])
    conflict_value["value_code"] = "VAL_LENGTH_CONFLICT"
    conflict_value["value"] = 4456
    conflict_value["resolution_state"] = "CONFLICTING"
    manifest["values"].append(conflict_value)
    manifest["conflict_decisions"] = [
        {
            "conflict_decision_code": "DEC_LENGTH_REVIEWED",
            "parameter_code": "overall_length_mm",
            "selected_value_code": "VAL_LENGTH_CONFLICT",
            "decision_state": "SELECTED",
            "rationale": "Explicit test decision selects the reviewed conflicting candidate.",
            "decided_at": "2026-08-31T00:00:00+07:00",
            "reviewer": "curation-test",
        }
    ]

    report = import_manifest(curated_session, manifest)
    assert report.axles == 2
    assert report.steering_relations == 1
    assert report.geometry_assets == 1
    assert report.conflicts == 1
    assert report.conflict_decisions == 1
    assert count(curated_session, ConflictDecision) == 1
