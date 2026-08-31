from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select

from app.db.models import VehicleConfiguration
from app.domain.enums import IdentityTimeBasis, IdentityVerificationState, EvidenceMethod
from app.domain.readiness import evaluate_readiness
from app.domain.schemas import (
    SourceDocumentCreate,
    SourceObservationCreate,
    VehicleConfigurationCreate,
)
from app.services.foundation import (
    create_normalized_value,
    create_source_document,
    create_source_observation,
    create_vehicle_configuration,
)
from app.domain.schemas import EvidenceLinkCreate, NormalizedValueCreate


FIXTURE_TIME = datetime(2026, 8, 31, tzinfo=timezone.utc)


def configuration_payload(**overrides) -> VehicleConfigurationCreate:
    values = {
        "stable_vehicle_code": "IDENTITY-TIME-TEST",
        "market_code": "TH",
        "generation_name": "Identity time test generation",
        "body_style": "hatchback",
        "variant_trim": "test edition",
    }
    values.update(overrides)
    return VehicleConfigurationCreate(**values)


def test_existing_exact_model_year_fixture_remains_valid(session):
    config = session.scalar(
        select(VehicleConfiguration).where(VehicleConfiguration.stable_vehicle_code == "FIXTURE-PRIMARY-PUBLISHED")
    )
    assert config.model_year_from == 2026
    assert config.identity_time_basis == IdentityTimeBasis.MODEL_YEAR.value


def test_exact_model_year_without_model_year_is_rejected():
    with pytest.raises(ValidationError, match="requires model_year_from"):
        configuration_payload(identity_time_basis=IdentityTimeBasis.MODEL_YEAR)


def test_exact_edition_release_can_have_null_model_year_and_preserves_raw_label():
    payload = configuration_payload(
        identity_time_basis=IdentityTimeBasis.EDITION_RELEASE,
        identity_time_label_raw="35th Anniversary Edition",
    )
    assert payload.model_year_from is None
    assert payload.identity_time_label_raw == "35th Anniversary Edition"


def test_exact_oem_revision_label_can_have_null_model_year_and_preserves_raw_label():
    payload = configuration_payload(
        identity_time_basis=IdentityTimeBasis.OEM_REVISION_LABEL,
        identity_time_label_raw="รุ่นปรับปรุงปี 2568",
    )
    assert payload.model_year_from is None
    assert payload.identity_time_label_raw == "รุ่นปรับปรุงปี 2568"


def test_exact_unknown_time_basis_is_rejected():
    with pytest.raises(ValidationError, match="cannot support RESOLVED_EXACT"):
        configuration_payload(identity_time_basis=IdentityTimeBasis.UNKNOWN)


def test_partial_unknown_time_basis_remains_representable():
    payload = configuration_payload(
        identity_verification_state=IdentityVerificationState.PARTIAL,
        identity_time_basis=IdentityTimeBasis.UNKNOWN,
    )
    assert payload.identity_verification_state == IdentityVerificationState.PARTIAL
    assert payload.identity_time_basis == IdentityTimeBasis.UNKNOWN


def test_exact_sale_period_requires_a_bounded_start():
    with pytest.raises(ValidationError, match="requires sale_period_from"):
        configuration_payload(identity_time_basis=IdentityTimeBasis.SALE_PERIOD)


def test_exact_multiple_requires_actual_temporal_evidence():
    with pytest.raises(ValidationError, match="MULTIPLE requires"):
        configuration_payload(identity_time_basis=IdentityTimeBasis.MULTIPLE)
    payload = configuration_payload(
        model_year_from=2024,
        identity_time_basis=IdentityTimeBasis.MULTIPLE,
        sale_period_from="2024-01-01",
    )
    assert payload.identity_time_basis == IdentityTimeBasis.MULTIPLE


def test_readiness_fails_closed_for_an_invalid_exact_identity_time_state(session):
    config = session.scalar(
        select(VehicleConfiguration).where(VehicleConfiguration.stable_vehicle_code == "FIXTURE-PRIMARY-PUBLISHED")
    )
    config.model_year_from = None
    config.identity_time_basis = IdentityTimeBasis.UNKNOWN.value
    identity = next(item for item in evaluate_readiness(session, config) if item.readiness_type.value == "IDENTITY_RESOLVED")
    assert identity.status.value == "NOT_READY"
    assert any("identity time" in reason for reason in identity.blocking_reasons)


def test_non_model_year_exact_record_roundtrips_searches_and_exports(client, session):
    config = create_vehicle_configuration(
        session,
        configuration_payload(
            stable_vehicle_code="IDENTITY-TIME-EDITION",
            identity_time_basis=IdentityTimeBasis.EDITION_RELEASE,
            identity_time_label_raw="35th Anniversary Edition",
        ),
        manufacturer_name="Identity Time Test Manufacturer",
        canonical_model_name="Identity Time Test Model",
        display_model_name="Identity Time Test Model",
    )
    source = create_source_document(
        session,
        SourceDocumentCreate(
            source_code="identity-time-test-source",
            title="Identity time deterministic source",
            publisher="Identity time test",
            authority_class="OEM_THAILAND",
            source_type="DETERMINISTIC_FIXTURE",
            market_code="TH",
            url="https://example.invalid/identity-time-test-source",
            retrieved_at=FIXTURE_TIME,
        ),
    )
    observation = create_source_observation(
        session,
        config,
        SourceObservationCreate(
            vehicle_identity_claim=config.stable_vehicle_code,
            source_document_id=source.id,
            raw_label="Overall length",
            raw_value="4500",
            raw_unit="mm",
            extracted_at=FIXTURE_TIME,
        ),
    )
    create_normalized_value(
        session,
        config,
        NormalizedValueCreate(
            parameter_code="overall_length_mm",
            numeric_value=4500,
            canonical_unit="mm",
            evidence_method=EvidenceMethod.PUBLISHED,
        ),
        evidence_links=[EvidenceLinkCreate(source_observation_id=observation.id)],
    )
    session.flush()

    search = client.get("/api/vehicles", params={"q": "35th Anniversary Edition"})
    assert search.status_code == 200
    assert search.json()["count"] == 1
    assert search.json()["items"][0]["identity_time_basis"] == "EDITION_RELEASE"
    assert search.json()["items"][0]["model_year_from"] is None

    detail = client.get("/api/vehicles/IDENTITY-TIME-EDITION")
    assert detail.status_code == 200
    body = detail.json()
    assert body["identity_time_basis"] == "EDITION_RELEASE"
    assert body["identity_time_label_raw"] == "35th Anniversary Edition"
    assert body["model_year_from"] is None
    html = client.get("/vehicles/IDENTITY-TIME-EDITION")
    assert "35th Anniversary Edition" in html.text

    csv_response = client.get("/exports/vehicles.csv", params={"codes": config.stable_vehicle_code})
    rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert rows[0]["identity_time_basis"] == "EDITION_RELEASE"
    assert rows[0]["identity_time_label_raw"] == "35th Anniversary Edition"
    assert rows[0]["model_year_from"] == ""

    xlsx_response = client.get("/exports/vehicles.xlsx", params={"codes": config.stable_vehicle_code})
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    sheet = workbook["Engineering Data"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    row = next(sheet.iter_rows(min_row=2))
    assert row[headers.index("identity_time_basis")].value == "EDITION_RELEASE"
    assert row[headers.index("identity_time_label_raw")].value == "35th Anniversary Edition"
    assert row[headers.index("model_year_from")].value is None
