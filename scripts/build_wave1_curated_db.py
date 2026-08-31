from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import joinedload, selectinload, sessionmaker

from app.curate.schemas import CurationManifest
from app.curate.validation import CurationError, load_manifest
from app.db.models import (
    Axle,
    ConflictDecision,
    EvidenceLink,
    GeometryAsset,
    LoadCondition,
    Manufacturer,
    NormalizedValue,
    ParameterAssessment,
    ParameterDefinition,
    ReadinessResult,
    SourceDocument,
    SourceObservation,
    SteeringRelation,
    VehicleConfiguration,
    VehicleFitment,
    VehicleModel,
)
from app.db.session import make_engine
from app.domain.readiness import evaluate_readiness
from app.exports.exporter import EXPORT_COLUMNS, csv_bytes, export_rows, xlsx_bytes


EXPECTED_MANIFEST_COUNT = 21
EXPECTED_SENTINEL_COUNT = 3
EXPECTED_WAVE1_COUNT = 18
READINESS_TYPES_PER_SCOPE = 4


class BuildError(RuntimeError):
    """Raised when the bounded Wave 1 build cannot pass an acceptance gate."""


@dataclass(frozen=True)
class ManifestInventory:
    paths: tuple[Path, ...]
    manifests: tuple[CurationManifest, ...]
    stable_vehicle_codes: tuple[str, ...]
    source_codes: tuple[str, ...]
    expected_counts: dict[str, int]


def _relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path)


def _resolved_path(path: Path) -> Path:
    return path if path.is_absolute() else ROOT / path


def _sqlite_url(path: Path) -> str:
    return f"sqlite:///{path.resolve().as_posix()}"


def _source_signature(source: Any) -> dict[str, Any]:
    return source.model_dump(mode="json")


def collect_inventory(root: Path = ROOT) -> ManifestInventory:
    """Load exactly the accepted sentinel and Wave 1 manifest set."""

    sentinel_paths = tuple(sorted((root / "data/curation/manifests/sentinel").glob("*.json")))
    wave1_paths = tuple(sorted((root / "data/curation/manifests/wave1").glob("*.json")))
    if len(sentinel_paths) != EXPECTED_SENTINEL_COUNT:
        raise BuildError(f"expected {EXPECTED_SENTINEL_COUNT} sentinel manifests, found {len(sentinel_paths)}")
    if len(wave1_paths) != EXPECTED_WAVE1_COUNT:
        raise BuildError(f"expected {EXPECTED_WAVE1_COUNT} Wave 1 manifests, found {len(wave1_paths)}")

    paths = sentinel_paths + wave1_paths
    try:
        manifests = tuple(load_manifest(path) for path in paths)
    except CurationError as exc:
        raise BuildError(str(exc)) from exc

    record_ids = [manifest.record_id for manifest in manifests]
    if len(record_ids) != len(set(record_ids)):
        raise BuildError("manifest set contains duplicate record_id values")
    stable_codes = [manifest.vehicle.stable_vehicle_code for manifest in manifests]
    if len(stable_codes) != len(set(stable_codes)):
        raise BuildError("manifest set contains duplicate stable_vehicle_code values")
    if len(stable_codes) != EXPECTED_MANIFEST_COUNT:
        raise BuildError(f"expected {EXPECTED_MANIFEST_COUNT} stable vehicle codes, found {len(stable_codes)}")

    source_signatures: dict[str, tuple[dict[str, Any], Path]] = {}
    for path, manifest in zip(paths, manifests):
        for source in manifest.sources:
            signature = _source_signature(source)
            previous = source_signatures.get(source.source_code)
            if previous is not None and previous[0] != signature:
                differing = sorted(
                    key for key in signature if signature[key] != previous[0].get(key)
                )
                raise BuildError(
                    f"source_code {source.source_code} has incompatible manifest metadata in "
                    f"{_relative(previous[1])} and {_relative(path)}: {', '.join(differing)}"
                )
            source_signatures[source.source_code] = (signature, path)

    expected_counts = {
        "vehicles": len(manifests),
        "source_entries": sum(len(manifest.sources) for manifest in manifests),
        "sources": len(source_signatures),
        "observations": sum(len(manifest.observations) for manifest in manifests),
        "values": sum(len(manifest.values) for manifest in manifests),
        "assessments": sum(len(manifest.assessments) for manifest in manifests),
        "loads": sum(len(manifest.load_conditions) for manifest in manifests),
        "fitments": sum(len(manifest.fitments) for manifest in manifests),
        "axles": sum(len(manifest.axles) for manifest in manifests),
        "steering_relations": sum(len(manifest.steering_relations) for manifest in manifests),
        "geometry_assets": sum(len(manifest.geometry_assets) for manifest in manifests),
        "conflict_decisions": sum(len(manifest.conflict_decisions) for manifest in manifests),
        "conflicting_values": sum(
            value.resolution_state.value == "CONFLICTING"
            for manifest in manifests
            for value in manifest.values
        ),
    }
    return ManifestInventory(
        paths=paths,
        manifests=manifests,
        stable_vehicle_codes=tuple(stable_codes),
        source_codes=tuple(sorted(source_signatures)),
        expected_counts=expected_counts,
    )


def _run_command(arguments: list[str], environment: dict[str, str]) -> None:
    command = [sys.executable, *arguments]
    display = " ".join(command)
    print(f"RUN {display}")
    try:
        subprocess.run(command, cwd=ROOT, env=environment, check=True)
    except subprocess.CalledProcessError as exc:
        raise BuildError(f"command failed with exit code {exc.returncode}: {display}") from exc


def _count(session: Any, model: Any) -> int:
    return int(session.scalar(select(func.count()).select_from(model)) or 0)


def _db_counts(session: Any) -> dict[str, int]:
    return {
        "vehicles": _count(session, VehicleConfiguration),
        "parameters": _count(session, ParameterDefinition),
        "manufacturers": _count(session, Manufacturer),
        "models": _count(session, VehicleModel),
        "sources": _count(session, SourceDocument),
        "observations": _count(session, SourceObservation),
        "values": _count(session, NormalizedValue),
        "assessments": _count(session, ParameterAssessment),
        "loads": _count(session, LoadCondition),
        "fitments": _count(session, VehicleFitment),
        "axles": _count(session, Axle),
        "steering_relations": _count(session, SteeringRelation),
        "geometry_assets": _count(session, GeometryAsset),
        "conflict_decisions": _count(session, ConflictDecision),
        "readiness_results": _count(session, ReadinessResult),
    }


def _registry_only_qa(database_url: str) -> dict[str, int]:
    """Prove that curate init populated only the parameter registry."""

    engine = make_engine(database_url)
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    try:
        with factory() as session:
            counts = _db_counts(session)
    finally:
        engine.dispose()

    if counts["parameters"] != 48:
        raise BuildError(
            f"registry-only init expected 48 parameter definitions, found {counts['parameters']}"
        )
    non_registry_rows = {
        key: value
        for key, value in counts.items()
        if key != "parameters" and value != 0
    }
    if non_registry_rows:
        raise BuildError(
            "registry-only init created non-registry rows: "
            + ", ".join(f"{key}={value}" for key, value in sorted(non_registry_rows.items()))
        )
    return {"parameters": counts["parameters"], "vehicles": counts["vehicles"]}


def _value_for(values: Iterable[NormalizedValue], config_id: str, parameter_code: str) -> list[NormalizedValue]:
    return [
        value
        for value in values
        if value.vehicle_configuration_id == config_id
        and value.parameter_definition
        and value.parameter_definition.parameter_code == parameter_code
    ]


def _assessment_for(
    assessments: Iterable[ParameterAssessment], config_id: str, parameter_code: str
) -> list[ParameterAssessment]:
    return [
        assessment
        for assessment in assessments
        if assessment.vehicle_configuration_id == config_id
        and assessment.parameter_definition
        and assessment.parameter_definition.parameter_code == parameter_code
    ]


def _qa_database(session: Any, inventory: ManifestInventory) -> dict[str, Any]:
    failures: list[str] = []

    def require(condition: bool, message: str) -> None:
        if not condition:
            failures.append(message)

    configs = list(
        session.scalars(
            select(VehicleConfiguration).options(selectinload(VehicleConfiguration.fitments))
        ).all()
    )
    values = list(
        session.scalars(
            select(NormalizedValue).options(
                joinedload(NormalizedValue.parameter_definition),
                joinedload(NormalizedValue.vehicle_fitment),
                joinedload(NormalizedValue.load_condition),
                selectinload(NormalizedValue.evidence_links)
                .joinedload(EvidenceLink.source_observation)
                .joinedload(SourceObservation.source_document),
            )
        ).all()
    )
    assessments = list(
        session.scalars(
            select(ParameterAssessment).options(
                joinedload(ParameterAssessment.parameter_definition),
                joinedload(ParameterAssessment.vehicle_fitment),
            )
        ).all()
    )
    observations = list(
        session.scalars(
            select(SourceObservation).options(
                joinedload(SourceObservation.source_document),
            )
        ).all()
    )
    readiness_rows = list(session.scalars(select(ReadinessResult)).all())
    conflict_decisions = list(session.scalars(select(ConflictDecision)).all())
    db_counts = _db_counts(session)

    for key in (
        "vehicles",
        "sources",
        "observations",
        "values",
        "assessments",
        "loads",
        "fitments",
        "axles",
        "steering_relations",
        "geometry_assets",
        "conflict_decisions",
    ):
        require(
            db_counts[key] == inventory.expected_counts[key],
            f"{key}: expected {inventory.expected_counts[key]}, found {db_counts[key]}",
        )
    require(db_counts["parameters"] == 48, f"parameters: expected 48, found {db_counts['parameters']}")
    require(
        db_counts["readiness_results"]
        == READINESS_TYPES_PER_SCOPE * (len(configs) + inventory.expected_counts["fitments"]),
        "readiness_results: one complete readiness set is required for every vehicle and fitment scope",
    )

    actual_codes = [config.stable_vehicle_code for config in configs]
    require(set(actual_codes) == set(inventory.stable_vehicle_codes), "database stable vehicle codes differ from manifests")
    require(len(actual_codes) == len(set(actual_codes)), "database stable vehicle codes are not unique")
    require(not any(code.startswith("FIXTURE-") for code in actual_codes), "Phase 0 synthetic vehicle rows were imported")

    for value in values:
        if value.evidence_method in {"PUBLISHED", "MEASURED"}:
            require(
                value.availability_state == "AVAILABLE",
                f"source-backed value {value.id} is not AVAILABLE",
            )
            require(bool(value.evidence_links), f"source-backed value {value.id} has no evidence link")
            for link in value.evidence_links:
                observation = link.source_observation
                require(observation is not None, f"value {value.id} has a missing source observation")
                if observation is not None:
                    require(
                        observation.vehicle_configuration_id == value.vehicle_configuration_id,
                        f"value {value.id} links to an observation from another vehicle",
                    )
                    require(
                        observation.source_document is not None,
                        f"observation {observation.id} has no source document",
                    )
        require(
            value.evidence_method not in {"DERIVED", "ESTIMATED"},
            f"direct {value.evidence_method} value {value.id} is not allowed in Wave 1 manifests",
        )
        if value.vehicle_fitment is not None:
            require(
                value.vehicle_fitment.vehicle_configuration_id == value.vehicle_configuration_id,
                f"value {value.id} references a fitment from another vehicle",
            )
        if value.load_condition is not None:
            require(
                value.load_condition.vehicle_configuration_id == value.vehicle_configuration_id,
                f"value {value.id} references a load condition from another vehicle",
            )

    for assessment in assessments:
        require(
            assessment.availability_state in {"UNKNOWN", "NOT_FOUND_AFTER_SEARCH", "NOT_APPLICABLE"},
            f"assessment {assessment.id} has invalid availability state {assessment.availability_state}",
        )
        if assessment.vehicle_fitment is not None:
            require(
                assessment.vehicle_fitment.vehicle_configuration_id == assessment.vehicle_configuration_id,
                f"assessment {assessment.id} references a fitment from another vehicle",
            )

    conflict_values = [
        value
        for value in values
        if value.resolution_state in {"CONFLICTING", "PREFERRED_WITH_CONFLICT"}
    ]
    require(
        len(conflict_values) == inventory.expected_counts["conflicting_values"],
        f"conflicting values: expected {inventory.expected_counts['conflicting_values']}, found {len(conflict_values)}",
    )
    require(
        len(conflict_decisions) == inventory.expected_counts["conflict_decisions"],
        f"conflict decisions: expected {inventory.expected_counts['conflict_decisions']}, found {len(conflict_decisions)}",
    )
    for value in conflict_values:
        require(bool(value.evidence_links), f"conflicting value {value.id} has no evidence lineage")

    by_code = {config.stable_vehicle_code: config for config in configs}
    readiness_statuses: Counter[tuple[str, str]] = Counter()
    readiness_by_key = {
        (row.vehicle_configuration_id, row.vehicle_fitment_id, row.readiness_type): row
        for row in readiness_rows
    }
    for config in configs:
        for fitment in [None, *config.fitments]:
            try:
                evaluations = evaluate_readiness(session, config, fitment=fitment)
            except Exception as exc:  # pragma: no cover - the message is part of the QA report
                failures.append(
                    f"readiness raised for {config.stable_vehicle_code}"
                    f"/{fitment.fitment_code if fitment else 'configuration'}: {exc}"
                )
                continue
            for evaluation in evaluations:
                status = evaluation.status.value
                readiness_statuses[(evaluation.readiness_type.value, status)] += 1
                row = readiness_by_key.get(
                    (config.id, fitment.id if fitment else None, evaluation.readiness_type.value)
                )
                require(
                    row is not None,
                    f"missing persisted readiness for {config.stable_vehicle_code}/{evaluation.readiness_type.value}",
                )
                if row is not None:
                    require(
                        row.status == status,
                        f"persisted readiness mismatch for {config.stable_vehicle_code}/{evaluation.readiness_type.value}",
                    )

    representative: dict[str, Any] = {}

    byd = by_code.get("th-byd-atto3-my24-extended-local")
    if byd is None:
        failures.append("missing BYD ATTO 3 representative")
    else:
        clearances = _value_for(values, byd.id, "clearance_value_mm")
        require({value.numeric_value for value in clearances} == {150.0, 175.0}, "BYD clearance values are not 150/175 mm")
        require(
            {value.load_condition.mass_basis for value in clearances if value.load_condition}
            == {"UNLADEN", "OEM_LADEN"},
            "BYD clearance load conditions are not UNLADEN/OEM_LADEN",
        )
        representative["BYD ATTO 3"] = {
            "clearance_mm": sorted(value.numeric_value for value in clearances),
            "clearance_load_basis": sorted(
                value.load_condition.mass_basis for value in clearances if value.load_condition
            ),
        }

    triton = by_code.get("th-mitsubishi-triton-ultra-4wd-at-release-2023")
    if triton is None:
        failures.append("missing Mitsubishi Triton representative")
    else:
        turning = _value_for(values, triton.id, "turning_radius_normalized_m")
        require(
            triton.model_year_from is None
            and triton.model_year_to is None
            and triton.identity_time_basis == "EDITION_RELEASE",
            "Mitsubishi Triton does not preserve the non-MY EDITION_RELEASE identity",
        )
        require(
            len(turning) == 1
            and turning[0].numeric_value == 6.2
            and turning[0].semantic_metadata
            == {
                "turning_radius_or_diameter": "RADIUS",
                "turning_reference": "OEM_UNSPECIFIED",
                "turning_axle_scope": "OEM_UNSPECIFIED",
            },
            "Mitsubishi Triton turning semantics are not preserved",
        )
        representative["Mitsubishi Triton"] = {
            "identity_time_basis": triton.identity_time_basis,
            "model_year_from": triton.model_year_from,
            "turning_radius_m": turning[0].numeric_value if turning else None,
            "turning_reference": turning[0].semantic_metadata.get("turning_reference") if turning else None,
        }

    volvo = by_code.get("th-volvo-ex30-ultra-smer-my2026-19")
    if volvo is None:
        failures.append("missing Volvo EX30 representative")
    else:
        volvo_fitment_codes = {fitment.fitment_code for fitment in volvo.fitments}
        width_codes = {
            "overall_width_body_mm",
            "overall_width_including_mirrors_mm",
            "overall_width_mirrors_folded_mm",
        }
        width_values = [value for value in values if value.vehicle_configuration_id == volvo.id and value.parameter_definition and value.parameter_definition.parameter_code in width_codes]
        turning_text = _value_for(values, volvo.id, "oem_turning_value_text")
        radius = _value_for(values, volvo.id, "turning_radius_normalized_m")
        volvo_decisions = [decision for decision in conflict_decisions if decision.vehicle_configuration_id == volvo.id]
        require(volvo_fitment_codes == {"WHEEL19"}, "Volvo EX30 WHEEL19 fitment scope is missing")
        require(
            {value.parameter_definition.parameter_code for value in width_values} == width_codes
            and all(value.vehicle_fitment_id is None for value in width_values),
            "Volvo EX30 body/mirrors-open/mirrors-folded widths are not distinct configuration values",
        )
        require(
            any(
                value.load_condition is not None
                and value.load_condition.name == "Kerb weight plus one person"
                for value in values
                if value.vehicle_configuration_id == volvo.id
                and value.parameter_definition
                and value.parameter_definition.parameter_code in {"overall_height_mm", "clearance_value_mm"}
            ),
            "Volvo EX30 kerb+1 load scope is missing",
        )
        require(
            len(turning_text) == 2
            and {value.text_value for value in turning_text} == {"10.7 m", "11 m"}
            and all(value.resolution_state == "CONFLICTING" for value in turning_text)
            and not radius
            and not volvo_decisions,
            "Volvo EX30 turning conflict/unknown semantics are not preserved",
        )
        representative["Volvo EX30"] = {
            "fitment_codes": sorted(volvo_fitment_codes),
            "turning_text_values": sorted(value.text_value for value in turning_text),
            "turning_radius_normalized": bool(radius),
            "conflict_decisions": len(volvo_decisions),
        }

    for code in (
        "th-honda-civic-ehev-rs-release-2026-07-23",
        "th-honda-accord-ehev-rs-release-2025-08-22",
    ):
        config = by_code.get(code)
        if config is None:
            failures.append(f"missing Honda representative {code}")
            continue
        b_values = [value for value in values if value.vehicle_configuration_id == config.id and value.authority_grade == "B"]
        require(bool(b_values), f"{code} has no authority-B values")
        require(
            all(
                {
                    link.source_observation.source_document.authority_class
                    for value in b_values
                    for link in value.evidence_links
                    if link.source_observation and link.source_observation.source_document
                }
                == {"REPUTABLE_SECONDARY"}
                for _ in [0]
            ),
            f"{code} authority-B values are not attributable only to REPUTABLE_SECONDARY evidence",
        )
        representative[code] = {"authority_B_values": len(b_values), "source_authority": "REPUTABLE_SECONDARY"}

    for code in (
        "th-tesla-model3-premium-long-range-rwd-2024plus",
        "th-tesla-modely-premium-lr-rwd-19-2025plus",
    ):
        config = by_code.get(code)
        if config is None:
            failures.append(f"missing Tesla representative {code}")
            continue
        text_values = _value_for(values, config.id, "oem_turning_value_text")
        radius = _value_for(values, config.id, "turning_radius_normalized_m")
        assessment = _assessment_for(assessments, config.id, "turning_radius_normalized_m")
        require(bool(text_values) and not radius and bool(assessment), f"{code} turning-circle raw text/radius assessment is not preserved")
        representative[code] = {"raw_turning_text": [value.text_value for value in text_values], "normalized_radius": bool(radius)}

    mg = by_code.get("th-mg-im6-long-range-release-2025-08-22")
    if mg is None:
        failures.append("missing MG IM6 representative")
    else:
        mg_observations = [observation for observation in observations if observation.vehicle_configuration_id == mg.id]
        rear_steer_observation = next(
            (
                observation
                for observation in mg_observations
                if "four-wheel steering" in observation.raw_value.lower()
            ),
            None,
        )
        mg_relations = [relation for relation in session.scalars(select(SteeringRelation)).all() if relation.vehicle_configuration_id == mg.id]
        require(rear_steer_observation is not None, "MG IM6 rear/four-wheel steering source observation is missing")
        require(not mg_relations, "MG IM6 has invented rear-steering kinematics")
        representative["MG IM6"] = {
            "four_wheel_steering_observation": rear_steer_observation.raw_value if rear_steer_observation else None,
            "steering_relations": len(mg_relations),
        }

    if failures:
        raise BuildError("database QA failed:\n" + "\n".join(f"- {failure}" for failure in failures))

    return {
        "database_counts": db_counts,
        "readiness_status": {
            f"{readiness_type}:{status}": count for (readiness_type, status), count in sorted(readiness_statuses.items())
        },
        "representative_proofs": representative,
        "source_reuse": {
            "manifest_entries": inventory.expected_counts["source_entries"],
            "unique_source_codes": inventory.expected_counts["sources"],
        },
    }


def _export_proof(
    session: Any,
    inventory: ManifestInventory,
    csv_path: Path,
    xlsx_path: Path,
) -> dict[str, Any]:
    rows = export_rows(session, inventory.stable_vehicle_codes)
    csv_data = csv_bytes(rows)
    xlsx_data = xlsx_bytes(rows)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.write_bytes(csv_data)
    xlsx_path.write_bytes(xlsx_data)

    csv_rows = list(csv.DictReader(csv_data.decode("utf-8-sig").splitlines()))
    workbook = load_workbook(BytesIO(xlsx_data), read_only=True)
    try:
        worksheet = workbook["Engineering Data"]
        headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        xlsx_rows = sum(1 for _ in worksheet.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    failures: list[str] = []
    expected_rows = inventory.expected_counts["values"] + inventory.expected_counts["assessments"]
    if len(csv_rows) != expected_rows:
        failures.append(f"CSV export row count: expected {expected_rows}, found {len(csv_rows)}")
    if xlsx_rows != expected_rows:
        failures.append(f"XLSX export row count: expected {expected_rows}, found {xlsx_rows}")
    if headers != EXPORT_COLUMNS:
        failures.append("XLSX headers do not match EXPORT_COLUMNS")
    if {row["stable_vehicle_code"] for row in csv_rows} != set(inventory.stable_vehicle_codes):
        failures.append("CSV export does not contain all 21 stable vehicle codes")
    if not all(
        row["value_kind"] == "assessment"
        or (row["source_observation_ids"] and row["source_document_codes"])
        for row in csv_rows
    ):
        failures.append("CSV export contains a source-backed row without observation/document provenance")
    if not any(row["fitment_code"] for row in csv_rows):
        failures.append("CSV export does not preserve any fitment scope")
    if not any(row["load_condition_id"] for row in csv_rows):
        failures.append("CSV export does not preserve any load-condition scope")
    if not any(row["resolution_state"] == "CONFLICTING" for row in csv_rows):
        failures.append("CSV export does not preserve conflict state")
    if b"password=" in csv_data.lower() or b"password=" in xlsx_data.lower():
        failures.append("export appears to contain a credential query field")
    if failures:
        raise BuildError("export proof failed:\n" + "\n".join(f"- {failure}" for failure in failures))

    return {
        "csv_path": _relative(csv_path),
        "xlsx_path": _relative(xlsx_path),
        "csv_rows": len(csv_rows),
        "vehicles": len({row["stable_vehicle_code"] for row in csv_rows}),
        "csv_bytes": len(csv_data),
        "xlsx_bytes": len(xlsx_data),
        "headers_match": headers == EXPORT_COLUMNS,
        "provenance_present": True,
        "fitment_scope_present": True,
        "load_scope_present": True,
        "conflict_state_present": True,
        "credentials_exposed": False,
    }


def build(args: argparse.Namespace) -> dict[str, Any]:
    inventory = collect_inventory()
    staging_path = _resolved_path(args.staging).resolve()
    final_path = _resolved_path(args.final).resolve()
    csv_path = _resolved_path(args.csv).resolve()
    xlsx_path = _resolved_path(args.xlsx).resolve()

    if staging_path == final_path:
        raise BuildError("staging and final database paths must differ")
    if staging_path.exists():
        raise BuildError(
            f"staging database already exists: {staging_path}; remove the failed/disposable staging file before retrying"
        )
    if not args.no_promote and final_path.exists() and not args.replace_final:
        raise BuildError(
            f"final database already exists: {final_path}; pass --replace-final only after a successful staging QA"
        )
    for sidecar in (staging_path.with_name(staging_path.name + "-wal"), staging_path.with_name(staging_path.name + "-shm")):
        if sidecar.exists():
            raise BuildError(f"staging sidecar already exists: {sidecar}")

    environment = os.environ.copy()
    environment["DATABASE_URL"] = _sqlite_url(staging_path)
    environment["PYTHONUNBUFFERED"] = "1"

    _run_command(["-m", "alembic", "upgrade", "head"], environment)
    _run_command(["-m", "app.curate", "init"], environment)
    registry_only = _registry_only_qa(environment["DATABASE_URL"])
    for manifest_path in inventory.paths:
        _run_command(["-m", "app.curate", "validate", str(manifest_path)], environment)
    for manifest_path in inventory.paths:
        _run_command(["-m", "app.curate", "import", str(manifest_path)], environment)

    engine = make_engine(_sqlite_url(staging_path))
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    try:
        with factory() as session:
            qa = _qa_database(session, inventory)
            exports = _export_proof(session, inventory, csv_path, xlsx_path)
    finally:
        engine.dispose()

    for sidecar in (staging_path.with_name(staging_path.name + "-wal"), staging_path.with_name(staging_path.name + "-shm")):
        if sidecar.exists():
            raise BuildError(f"refusing to promote while SQLite sidecar exists: {sidecar}")

    promoted_to: str | None = None
    if not args.no_promote:
        final_path.parent.mkdir(parents=True, exist_ok=True)
        staging_path.replace(final_path)
        promoted_to = _relative(final_path)

    return {
        "status": "PASS",
        "manifest_count": len(inventory.manifests),
        "sentinel_count": EXPECTED_SENTINEL_COUNT,
        "wave1_count": EXPECTED_WAVE1_COUNT,
        "stable_vehicle_codes": list(inventory.stable_vehicle_codes),
        "expected_manifest_counts": inventory.expected_counts,
        "staging_database": _relative(staging_path),
        "promoted_database": promoted_to,
        "registry_only": registry_only,
        **qa,
        "exports": exports,
    }


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build the accepted Wave 1 curated SQLite database")
    parser.add_argument(
        "--staging",
        type=Path,
        default=Path("vehicle_engineering_curated.staging.db"),
        help="clean disposable staging database path (default: vehicle_engineering_curated.staging.db)",
    )
    parser.add_argument(
        "--final",
        type=Path,
        default=Path("vehicle_engineering_curated.db"),
        help="accepted local database path (default: vehicle_engineering_curated.db)",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=Path("vehicle_engineering_curated.wave1.csv"),
        help="ignored CSV proof output path",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("vehicle_engineering_curated.wave1.xlsx"),
        help="ignored XLSX proof output path",
    )
    parser.add_argument(
        "--no-promote",
        action="store_true",
        help="leave the successful staging DB in place for inspection",
    )
    parser.add_argument(
        "--replace-final",
        action="store_true",
        help="allow promotion to replace an existing final database after all gates pass",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    try:
        report = build(_parser().parse_args(argv))
    except (BuildError, OSError, ValueError) as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
