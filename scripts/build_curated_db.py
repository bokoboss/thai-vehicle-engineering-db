"""Build and safely promote a curated database from an accepted release set.

The release definition is the production membership boundary.  This script
never discovers manifests by scanning a directory, and it never writes to the
accepted database until a disposable staging database has passed every gate.
"""

from __future__ import annotations

import argparse
import csv
from contextlib import contextmanager
import hashlib
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
import uuid
from collections import Counter
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping

try:
    import fcntl
except ImportError:  # pragma: no cover - Windows provides msvcrt instead.
    fcntl = None

try:
    import msvcrt
except ImportError:  # pragma: no cover - POSIX provides fcntl instead.
    msvcrt = None


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import joinedload, selectinload, sessionmaker

from app.curate.schemas import CurationManifest
from app.curate.validation import CurationError, load_manifest
from app.db.models import (
    Axle,
    ConflictDecision,
    EvidenceLink,
    GeometryAsset,
    LoadCondition,
    Manufacturer,
    NormalizedValue,
    ParameterAssessment,
    ParameterDefinition,
    ReadinessResult,
    SourceDocument,
    SourceObservation,
    SteeringRelation,
    VehicleConfiguration,
    VehicleFitment,
    VehicleModel,
)
from app.db.session import make_engine
from app.domain.readiness import evaluate_readiness
from app.exports.exporter import EXPORT_COLUMNS, csv_bytes, export_rows, xlsx_bytes
from app.seed.registry import load_registry


CURRENT_RELEASE_POINTER_PATH = Path("data/curation/releases/current_release.json")
CURRENT_RELEASE_POINTER_SCHEMA_VERSION = "1.0"
# Compatibility name for callers that imported the old default constant.  The
# value is now the stable selector, never a versioned release definition.
DEFAULT_RELEASE_PATH = CURRENT_RELEASE_POINTER_PATH
READINESS_TYPES_PER_SCOPE = 4
RELEASE_SCHEMA_VERSION = "1.0"
BUILD_INPUT_FINGERPRINT_SCHEMA_VERSION = "1.0"
BUILD_COMPATIBILITY_VERSION = "2.0"
PROMOTED_DATABASE_METADATA_SCHEMA_VERSION = "1.0"
PARAMETER_REGISTRY_RELATIVE_PATH = Path("data/reference/parameter_registry_v1.json")
MIGRATION_RELATIVE_DIRECTORY = Path("alembic/versions")
REFRESH_LOCK_SUFFIX = ".refresh.lock"
REFRESH_LOCK_TIMEOUT_SECONDS = 30.0
REFRESH_LOCK_POLL_SECONDS = 0.05
# Explicitly enumerate build/schema code that can change the produced DB.
# Hashing content plus repository-relative labels keeps the result portable.
BUILD_COMPATIBILITY_FILES = (
    Path("alembic.ini"),
    Path("alembic/env.py"),
    Path("app/config.py"),
    Path("app/curate/__main__.py"),
    Path("app/curate/loader.py"),
    Path("app/curate/report.py"),
    Path("app/curate/schemas.py"),
    Path("app/curate/validation.py"),
    Path("app/db/base.py"),
    Path("app/db/models/__init__.py"),
    Path("app/db/models/entities.py"),
    Path("app/db/session.py"),
    Path("app/domain/candidate_resolution.py"),
    Path("app/domain/enums.py"),
    Path("app/domain/readiness.py"),
    Path("app/domain/schemas.py"),
    Path("app/domain/scope.py"),
    Path("app/domain/validation.py"),
    Path("app/exports/exporter.py"),
    Path("app/seed/registry.py"),
    Path("app/services/foundation.py"),
    Path("scripts/build_curated_db.py"),
)
_RELEASE_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]*$")
_VERSIONED_RELEASE_FILENAME_RE = re.compile(
    r"^release_[A-Za-z0-9][A-Za-z0-9_.-]*\.json$"
)


class BuildError(RuntimeError):
    """Raised when a curated release cannot pass an acceptance gate."""


def refresh_lock_path(database_path: Path) -> Path:
    """Return the shared advisory-lock path for one accepted database."""

    database_path = Path(database_path).resolve()
    return database_path.with_name(database_path.name + REFRESH_LOCK_SUFFIX)


@contextmanager
def database_refresh_lock(
    database_path: Path,
    *,
    timeout_seconds: float = REFRESH_LOCK_TIMEOUT_SECONDS,
    poll_seconds: float = REFRESH_LOCK_POLL_SECONDS,
) -> Iterator[None]:
    """Serialize refresh/promotion of one accepted DB across processes.

    The adjacent file is only the stable OS-lock target; its existence is not
    used as a sentinel.  The operating system releases the advisory lock when
    the owning file handle or process exits.
    """

    if timeout_seconds < 0 or poll_seconds <= 0:
        raise ValueError("lock timeout must be non-negative and poll interval must be positive")
    if fcntl is None and msvcrt is None:
        raise BuildError("curated database refresh locking is unsupported on this platform")

    lock_path = refresh_lock_path(database_path)
    try:
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        handle = lock_path.open("a+b")
    except OSError as exc:
        raise BuildError(f"could not open curated database refresh lock {lock_path}: {exc}") from exc

    acquired = False
    deadline = time.monotonic() + timeout_seconds
    try:
        if handle.seek(0, os.SEEK_END) == 0:
            handle.write(b"\0")
            handle.flush()
        while not acquired:
            try:
                if fcntl is not None:
                    fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                else:
                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                acquired = True
            except (BlockingIOError, OSError) as exc:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    raise BuildError(
                        "timed out acquiring curated database refresh lock "
                        f"{lock_path}; another launcher or build may still be refreshing"
                    ) from exc
                time.sleep(min(poll_seconds, remaining))
        yield
    finally:
        if acquired:
            try:
                if fcntl is not None:
                    fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
                else:
                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            finally:
                handle.close()
        else:
            handle.close()


@dataclass(frozen=True)
class ReleaseDefinition:
    schema_version: str
    release_id: str
    release_date: str
    release_status: str
    manifest_paths: tuple[str, ...]
    data_standard_version: str | None = None
    methodology_versions: Mapping[str, str] = field(default_factory=dict)
    notes: str | None = None
    path: Path | None = None


@dataclass(frozen=True)
class ManifestInventory:
    release: ReleaseDefinition
    paths: tuple[Path, ...]
    manifests: tuple[CurationManifest, ...]
    stable_vehicle_codes: tuple[str, ...]
    source_codes: tuple[str, ...]
    expected_counts: dict[str, int]
    stable_vehicle_digest: str
    build_input_digest_sha256: str


def _relative(path: Path, root: Path = ROOT) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return Path(path).as_posix()


def _portable_relative(path: Path | None, root: Path = ROOT) -> str | None:
    """Report repository paths without leaking disposable absolute paths."""

    if path is None:
        return None
    path = Path(path).resolve()
    try:
        return path.relative_to(root.resolve()).as_posix()
    except ValueError:
        return f"external/{path.name}"


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _reject_json_constant(value: str) -> Any:
    raise ValueError(f"non-standard JSON constant is not allowed: {value}")


def _load_json_value(path: Path) -> Any:
    try:
        return json.loads(
            path.read_text(encoding="utf-8"),
            object_pairs_hook=_duplicate_key_guard,
            parse_constant=_reject_json_constant,
        )
    except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise BuildError(f"invalid JSON input {path}: {exc}") from exc


def _resolve_fingerprint_input(root: Path, relative_path: Path) -> Path:
    candidate = (root / relative_path).resolve()
    if candidate.is_file():
        return candidate
    fallback = (ROOT / relative_path).resolve()
    if fallback.is_file():
        return fallback
    raise BuildError(f"build compatibility input was not found: {relative_path.as_posix()}")


def _text_input_sha256(path: Path) -> str:
    normalized = path.read_text(encoding="utf-8").replace("\r\n", "\n").replace("\r", "\n")
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _migration_compatibility_files(root: Path) -> tuple[tuple[Path, Path], ...]:
    """Return all migration files with stable repository-relative identities."""

    versions_dir = (root / MIGRATION_RELATIVE_DIRECTORY).resolve()
    if not versions_dir.is_dir():
        versions_dir = (ROOT / MIGRATION_RELATIVE_DIRECTORY).resolve()
    files = [
        (
            MIGRATION_RELATIVE_DIRECTORY / path.relative_to(versions_dir),
            path,
        )
        for path in versions_dir.glob("*.py")
        if path.is_file()
    ]
    return tuple(sorted(files, key=lambda item: item[0].as_posix()))


def _build_compatibility_payload(root: Path) -> dict[str, Any]:
    compatibility_inputs = [
        (
            relative_path,
            _resolve_fingerprint_input(root, relative_path),
        )
        for relative_path in BUILD_COMPATIBILITY_FILES
    ]
    compatibility_inputs.extend(_migration_compatibility_files(root))
    compatibility_inputs.sort(key=lambda item: item[0].as_posix())
    return {
        "compatibility_version": BUILD_COMPATIBILITY_VERSION,
        "release_schema_version": RELEASE_SCHEMA_VERSION,
        "current_release_pointer_schema_version": CURRENT_RELEASE_POINTER_SCHEMA_VERSION,
        "metadata_schema_version": PROMOTED_DATABASE_METADATA_SCHEMA_VERSION,
        "inputs": [
            {
                "path": relative_path.as_posix(),
                "sha256": _text_input_sha256(path),
            }
            for relative_path, path in compatibility_inputs
        ],
    }


def build_input_fingerprint(
    inventory: ManifestInventory,
    *,
    root: Path = ROOT,
    compatibility_version: str | None = None,
) -> str:
    """Hash the exact accepted-release inputs used by the curated build."""

    root = root.resolve()
    if inventory.release.path is None:
        raise BuildError("cannot fingerprint a release without its definition path")

    manifests = [
        {
            "path": Path(relative_path).as_posix(),
            "content": _load_json_value(manifest_path),
        }
        for relative_path, manifest_path in zip(
            inventory.release.manifest_paths,
            inventory.paths,
        )
    ]
    manifests.sort(key=lambda item: item["path"])

    compatibility = _build_compatibility_payload(root)
    if compatibility_version is not None:
        compatibility["compatibility_version"] = compatibility_version

    payload = {
        "fingerprint_schema_version": BUILD_INPUT_FINGERPRINT_SCHEMA_VERSION,
        "release_definition": {
            "path": _portable_relative(inventory.release.path, root),
            "content": _load_json_value(inventory.release.path),
        },
        "manifests": manifests,
        "parameter_registry": {
            "path": PARAMETER_REGISTRY_RELATIVE_PATH.as_posix(),
            "content": _load_json_value(
                _resolve_fingerprint_input(root, PARAMETER_REGISTRY_RELATIVE_PATH)
            ),
        },
        "build_compatibility": compatibility,
    }
    return hashlib.sha256(_canonical_json_bytes(payload)).hexdigest()


def _resolved_path(path: Path, root: Path = ROOT) -> Path:
    return path if path.is_absolute() else root / path


def _sqlite_url(path: Path) -> str:
    return f"sqlite:///{path.resolve().as_posix()}"


def _duplicate_key_guard(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON field: {key}")
        result[key] = value
    return result


def _load_json_object(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_text(encoding="utf-8")
        value = json.loads(raw, object_pairs_hook=_duplicate_key_guard)
    except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise BuildError(f"invalid release JSON {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise BuildError(f"release definition root must be an object: {path}")
    return value


def _load_release_definition_file(path: Path, *, root: Path) -> ReleaseDefinition:
    """Load and validate one explicit, versioned release definition file."""

    path = path.resolve()
    if not path.is_file():
        raise BuildError(f"release definition was not found: {_relative(path, root)}")
    document = _load_json_object(path)
    allowed = {
        "release_schema_version",
        "release_id",
        "release_date",
        "release_status",
        "manifest_paths",
        "data_standard_version",
        "methodology_versions",
        "notes",
    }
    unknown = sorted(set(document) - allowed)
    if unknown:
        raise BuildError(f"release definition has unknown fields: {', '.join(unknown)}")

    required = {
        "release_schema_version",
        "release_id",
        "release_date",
        "release_status",
        "manifest_paths",
    }
    missing = sorted(required - set(document))
    if missing:
        raise BuildError(f"release definition is missing fields: {', '.join(missing)}")

    schema_version = document["release_schema_version"]
    if schema_version != RELEASE_SCHEMA_VERSION:
        raise BuildError(
            f"unsupported release schema version {schema_version!r}; expected {RELEASE_SCHEMA_VERSION!r}"
        )

    release_id = document["release_id"]
    if not isinstance(release_id, str) or not _RELEASE_ID_RE.fullmatch(release_id):
        raise BuildError("release_id must be a simple non-empty identifier")

    release_date = document["release_date"]
    if not isinstance(release_date, str):
        raise BuildError("release_date must be an ISO date string")
    try:
        date.fromisoformat(release_date)
    except ValueError as exc:
        raise BuildError(f"release_date is not an ISO date: {release_date!r}") from exc

    release_status = document["release_status"]
    if release_status != "ACCEPTED":
        raise BuildError(f"release_status must be ACCEPTED, found {release_status!r}")

    raw_paths = document["manifest_paths"]
    if not isinstance(raw_paths, list) or not raw_paths:
        raise BuildError("manifest_paths must be a non-empty JSON array")

    root = root.resolve()
    normalized_paths: list[str] = []
    seen_paths: set[str] = set()
    for raw_manifest_path in raw_paths:
        if not isinstance(raw_manifest_path, str) or not raw_manifest_path.strip():
            raise BuildError("manifest_paths entries must be non-empty strings")
        if "\\" in raw_manifest_path:
            raise BuildError(
                f"manifest path must use repository-relative '/' separators: {raw_manifest_path!r}"
            )
        candidate = Path(raw_manifest_path)
        if candidate.is_absolute():
            raise BuildError(f"manifest path must be repository-relative: {raw_manifest_path!r}")
        resolved = (root / candidate).resolve()
        try:
            relative = resolved.relative_to(root)
        except ValueError as exc:
            raise BuildError(f"manifest path escapes repository root: {raw_manifest_path!r}") from exc
        canonical = relative.as_posix()
        key = canonical.casefold()
        if key in seen_paths:
            raise BuildError(f"release contains duplicate manifest path: {canonical}")
        seen_paths.add(key)
        if resolved.suffix.lower() != ".json":
            raise BuildError(f"listed manifest is not JSON: {canonical}")
        if not resolved.is_file():
            raise BuildError(f"listed manifest is missing: {canonical}")
        normalized_paths.append(canonical)

    data_standard_version = document.get("data_standard_version")
    if data_standard_version is not None and not isinstance(data_standard_version, str):
        raise BuildError("data_standard_version must be a string when present")
    methodology_versions = document.get("methodology_versions", {})
    if not isinstance(methodology_versions, dict) or not all(
        isinstance(key, str) and isinstance(value, str)
        for key, value in methodology_versions.items()
    ):
        raise BuildError("methodology_versions must be an object of string values")
    notes = document.get("notes")
    if notes is not None and not isinstance(notes, str):
        raise BuildError("release notes must be a string when present")

    return ReleaseDefinition(
        schema_version=schema_version,
        release_id=release_id,
        release_date=release_date,
        release_status=release_status,
        manifest_paths=tuple(normalized_paths),
        data_standard_version=data_standard_version,
        methodology_versions=dict(methodology_versions),
        notes=notes,
        path=path,
    )


def resolve_current_release_path(*, root: Path = ROOT) -> Path:
    """Resolve and validate the repository's stable current-release pointer."""

    root = root.resolve()
    pointer_path = (root / CURRENT_RELEASE_POINTER_PATH).resolve()
    if not pointer_path.is_file():
        raise BuildError(
            f"current release pointer was not found: {_relative(pointer_path, root)}"
        )

    try:
        document = _load_json_object(pointer_path)
    except BuildError as exc:
        raise BuildError(
            f"invalid current release pointer {_relative(pointer_path, root)}: {exc}"
        ) from exc

    allowed = {"pointer_schema_version", "release"}
    unknown = sorted(set(document) - allowed)
    if unknown:
        raise BuildError(
            "current release pointer has unknown fields: " + ", ".join(unknown)
        )
    missing = sorted(allowed - set(document))
    if missing:
        raise BuildError(
            "current release pointer is missing fields: " + ", ".join(missing)
        )

    schema_version = document["pointer_schema_version"]
    if schema_version != CURRENT_RELEASE_POINTER_SCHEMA_VERSION:
        raise BuildError(
            "unsupported current release pointer schema version "
            f"{schema_version!r}; expected {CURRENT_RELEASE_POINTER_SCHEMA_VERSION!r}"
        )

    target_name = document["release"]
    if not isinstance(target_name, str) or not target_name or target_name != target_name.strip():
        raise BuildError(
            "current release pointer release must be a non-empty filename string"
        )
    if "\\" in target_name:
        raise BuildError(
            "current release pointer target must use a direct filename with '/' semantics"
        )

    target = Path(target_name)
    releases_dir = (root / CURRENT_RELEASE_POINTER_PATH).parent.resolve()
    if target.is_absolute() or target.parent != Path("."):
        raise BuildError(
            "current release pointer target must be a versioned release filename "
            "inside data/curation/releases"
        )
    if target.name.casefold() == pointer_path.name.casefold():
        raise BuildError("current release pointer cannot target itself")
    if not _VERSIONED_RELEASE_FILENAME_RE.fullmatch(target_name):
        raise BuildError(
            "current release pointer target must match release_*.json"
        )

    target_path = (releases_dir / target).resolve()
    try:
        target_path.relative_to(releases_dir)
    except ValueError as exc:
        raise BuildError(
            "current release pointer target escapes data/curation/releases"
        ) from exc
    if not target_path.is_file():
        raise BuildError(
            "current release pointer target was not found: "
            f"{_relative(target_path, root)}"
        )

    # Validate the target itself before returning it.  This keeps an accepted
    # pointer from bypassing the existing release-definition contract.
    _load_release_definition_file(target_path, root=root)
    return target_path


def resolve_release_path(
    release_path: str | Path | None = None,
    *,
    root: Path = ROOT,
) -> Path:
    """Resolve an explicit release or the validated current-release pointer."""

    root = root.resolve()
    if release_path is None:
        return resolve_current_release_path(root=root)

    path = _resolved_path(Path(release_path), root).resolve()
    pointer_path = (root / CURRENT_RELEASE_POINTER_PATH).resolve()
    if path == pointer_path:
        return resolve_current_release_path(root=root)
    return path


def load_release_definition(
    release_path: str | Path | None = None,
    *,
    root: Path = ROOT,
) -> ReleaseDefinition:
    """Load an explicit release or the validated current-release target."""

    return _load_release_definition_file(
        resolve_release_path(release_path, root=root),
        root=root,
    )


def _source_signature(source: Any) -> dict[str, Any]:
    return source.model_dump(mode="json")


def collect_inventory(
    release_path: str | Path | None = None,
    *,
    root: Path = ROOT,
) -> ManifestInventory:
    """Load exactly the manifests explicitly listed by an accepted release."""

    root = root.resolve()
    release = load_release_definition(release_path, root=root)
    paths = tuple((root / relative_path).resolve() for relative_path in release.manifest_paths)
    try:
        manifests = tuple(load_manifest(path) for path in paths)
    except CurationError as exc:
        raise BuildError(str(exc)) from exc

    record_ids = [manifest.record_id for manifest in manifests]
    if len(record_ids) != len(set(record_ids)):
        raise BuildError("release contains duplicate record_id values")
    stable_codes = [manifest.vehicle.stable_vehicle_code for manifest in manifests]
    if len(stable_codes) != len(set(stable_codes)):
        raise BuildError("release contains duplicate stable_vehicle_code values")
    fixture_codes = sorted(code for code in stable_codes if code.startswith("FIXTURE-"))
    if fixture_codes:
        raise BuildError(
            "release contains Phase 0 fixture identities: " + ", ".join(fixture_codes)
        )

    source_signatures: dict[str, tuple[dict[str, Any], Path]] = {}
    for path, manifest in zip(paths, manifests):
        for source in manifest.sources:
            signature = _source_signature(source)
            previous = source_signatures.get(source.source_code)
            if previous is not None and previous[0] != signature:
                differing = sorted(
                    key for key in signature if signature[key] != previous[0].get(key)
                )
                raise BuildError(
                    f"source_code {source.source_code} has incompatible manifest metadata in "
                    f"{_relative(previous[1], root)} and {_relative(path, root)}: "
                    f"{', '.join(differing)}"
                )
            source_signatures[source.source_code] = (signature, path)

    expected_counts = {
        "vehicles": len(manifests),
        "source_entries": sum(len(manifest.sources) for manifest in manifests),
        "sources": len(source_signatures),
        "observations": sum(len(manifest.observations) for manifest in manifests),
        "values": sum(len(manifest.values) for manifest in manifests),
        "assessments": sum(len(manifest.assessments) for manifest in manifests),
        "loads": sum(len(manifest.load_conditions) for manifest in manifests),
        "fitments": sum(len(manifest.fitments) for manifest in manifests),
        "axles": sum(len(manifest.axles) for manifest in manifests),
        "steering_relations": sum(len(manifest.steering_relations) for manifest in manifests),
        "geometry_assets": sum(len(manifest.geometry_assets) for manifest in manifests),
        "conflict_decisions": sum(len(manifest.conflict_decisions) for manifest in manifests),
        "conflicting_values": sum(
            value.resolution_state.value in {"CONFLICTING", "PREFERRED_WITH_CONFLICT"}
            for manifest in manifests
            for value in manifest.values
        ),
    }
    stable_vehicle_digest = hashlib.sha256(
        "\n".join(sorted(stable_codes)).encode("utf-8")
    ).hexdigest()
    inventory = ManifestInventory(
        release=release,
        paths=paths,
        manifests=manifests,
        stable_vehicle_codes=tuple(stable_codes),
        source_codes=tuple(sorted(source_signatures)),
        expected_counts=expected_counts,
        stable_vehicle_digest=stable_vehicle_digest,
        build_input_digest_sha256="",
    )
    return replace(
        inventory,
        build_input_digest_sha256=build_input_fingerprint(inventory, root=root),
    )


def _run_command(
    arguments: list[str],
    environment: dict[str, str],
    *,
    root: Path = ROOT,
) -> None:
    command = [sys.executable, *arguments]
    display = " ".join(command)
    print(f"RUN {display}")
    try:
        subprocess.run(command, cwd=root, env=environment, check=True)
    except subprocess.CalledProcessError as exc:
        raise BuildError(f"command failed with exit code {exc.returncode}: {display}") from exc


def _count(session: Any, model: Any) -> int:
    return int(session.scalar(select(func.count()).select_from(model)) or 0)


def _db_counts(session: Any) -> dict[str, int]:
    return {
        "vehicles": _count(session, VehicleConfiguration),
        "parameters": _count(session, ParameterDefinition),
        "manufacturers": _count(session, Manufacturer),
        "models": _count(session, VehicleModel),
        "sources": _count(session, SourceDocument),
        "observations": _count(session, SourceObservation),
        "values": _count(session, NormalizedValue),
        "assessments": _count(session, ParameterAssessment),
        "loads": _count(session, LoadCondition),
        "fitments": _count(session, VehicleFitment),
        "axles": _count(session, Axle),
        "steering_relations": _count(session, SteeringRelation),
        "geometry_assets": _count(session, GeometryAsset),
        "conflict_decisions": _count(session, ConflictDecision),
        "readiness_results": _count(session, ReadinessResult),
    }


def _registry_only_qa(database_url: str) -> dict[str, int]:
    """Prove that curate init populated only the accepted parameter registry."""

    engine = make_engine(database_url)
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    try:
        with factory() as session:
            counts = _db_counts(session)
    finally:
        engine.dispose()

    expected_parameters = len(load_registry().parameters)
    if counts["parameters"] != expected_parameters:
        raise BuildError(
            f"registry-only init expected {expected_parameters} parameter definitions, "
            f"found {counts['parameters']}"
        )
    non_registry_rows = {
        key: value
        for key, value in counts.items()
        if key != "parameters" and value != 0
    }
    if non_registry_rows:
        raise BuildError(
            "registry-only init created non-registry rows: "
            + ", ".join(f"{key}={value}" for key, value in sorted(non_registry_rows.items()))
        )
    return {"parameters": counts["parameters"], "vehicles": counts["vehicles"]}


def _value_for(
    values: Iterable[NormalizedValue],
    config_id: str,
    parameter_code: str,
) -> list[NormalizedValue]:
    return [
        value
        for value in values
        if value.vehicle_configuration_id == config_id
        and value.parameter_definition
        and value.parameter_definition.parameter_code == parameter_code
    ]


def _assessment_for(
    assessments: Iterable[ParameterAssessment],
    config_id: str,
    parameter_code: str,
) -> list[ParameterAssessment]:
    return [
        assessment
        for assessment in assessments
        if assessment.vehicle_configuration_id == config_id
        and assessment.parameter_definition
        and assessment.parameter_definition.parameter_code == parameter_code
    ]


def _qa_database(session: Any, inventory: ManifestInventory) -> dict[str, Any]:
    failures: list[str] = []

    def require(condition: bool, message: str) -> None:
        if not condition:
            failures.append(message)

    configs = list(
        session.scalars(
            select(VehicleConfiguration).options(selectinload(VehicleConfiguration.fitments))
        ).all()
    )
    values = list(
        session.scalars(
            select(NormalizedValue).options(
                joinedload(NormalizedValue.parameter_definition),
                joinedload(NormalizedValue.vehicle_fitment),
                joinedload(NormalizedValue.load_condition),
                selectinload(NormalizedValue.evidence_links)
                .joinedload(EvidenceLink.source_observation)
                .joinedload(SourceObservation.source_document),
            )
        ).all()
    )
    assessments = list(
        session.scalars(
            select(ParameterAssessment).options(
                joinedload(ParameterAssessment.parameter_definition),
                joinedload(ParameterAssessment.vehicle_fitment),
            )
        ).all()
    )
    observations = list(
        session.scalars(
            select(SourceObservation).options(joinedload(SourceObservation.source_document))
        ).all()
    )
    readiness_rows = list(session.scalars(select(ReadinessResult)).all())
    conflict_decisions = list(session.scalars(select(ConflictDecision)).all())
    db_counts = _db_counts(session)
    expected_registry_count = len(load_registry().parameters)

    for key in (
        "vehicles",
        "sources",
        "observations",
        "values",
        "assessments",
        "loads",
        "fitments",
        "axles",
        "steering_relations",
        "geometry_assets",
        "conflict_decisions",
    ):
        require(
            db_counts[key] == inventory.expected_counts[key],
            f"{key}: expected {inventory.expected_counts[key]}, found {db_counts[key]}",
        )
    require(
        db_counts["parameters"] == expected_registry_count,
        f"parameters: expected {expected_registry_count}, found {db_counts['parameters']}",
    )
    require(
        db_counts["readiness_results"]
        == READINESS_TYPES_PER_SCOPE * (len(configs) + inventory.expected_counts["fitments"]),
        "readiness_results: one complete readiness set is required for every vehicle and fitment scope",
    )

    actual_codes = [config.stable_vehicle_code for config in configs]
    require(
        set(actual_codes) == set(inventory.stable_vehicle_codes),
        "database stable vehicle codes differ from manifests",
    )
    require(len(actual_codes) == len(set(actual_codes)), "database stable vehicle codes are not unique")
    require(not any(code.startswith("FIXTURE-") for code in actual_codes), "Phase 0 synthetic rows were imported")
    actual_source_codes = {
        source.source_code
        for source in session.scalars(select(SourceDocument)).all()
    }
    require(actual_source_codes == set(inventory.source_codes), "database source codes differ from release manifests")

    for value in values:
        if value.evidence_method in {"PUBLISHED", "MEASURED"}:
            require(value.availability_state == "AVAILABLE", f"source-backed value {value.id} is not AVAILABLE")
            require(bool(value.evidence_links), f"source-backed value {value.id} has no evidence link")
            for link in value.evidence_links:
                observation = link.source_observation
                require(observation is not None, f"value {value.id} has a missing source observation")
                if observation is not None:
                    require(
                        observation.vehicle_configuration_id == value.vehicle_configuration_id,
                        f"value {value.id} links to an observation from another vehicle",
                    )
                    require(observation.source_document is not None, f"observation {observation.id} has no source document")
        require(
            value.evidence_method not in {"DERIVED", "ESTIMATED"},
            f"direct {value.evidence_method} value {value.id} is not allowed in release manifests",
        )
        if value.vehicle_fitment is not None:
            require(
                value.vehicle_fitment.vehicle_configuration_id == value.vehicle_configuration_id,
                f"value {value.id} references a fitment from another vehicle",
            )
        if value.load_condition is not None:
            require(
                value.load_condition.vehicle_configuration_id == value.vehicle_configuration_id,
                f"value {value.id} references a load condition from another vehicle",
            )

    for assessment in assessments:
        require(
            assessment.availability_state in {"UNKNOWN", "NOT_FOUND_AFTER_SEARCH", "NOT_APPLICABLE"},
            f"assessment {assessment.id} has invalid availability state {assessment.availability_state}",
        )
        if assessment.vehicle_fitment is not None:
            require(
                assessment.vehicle_fitment.vehicle_configuration_id == assessment.vehicle_configuration_id,
                f"assessment {assessment.id} references a fitment from another vehicle",
            )

    conflict_values = [
        value
        for value in values
        if value.resolution_state in {"CONFLICTING", "PREFERRED_WITH_CONFLICT"}
    ]
    require(
        len(conflict_values) == inventory.expected_counts["conflicting_values"],
        f"conflicting values: expected {inventory.expected_counts['conflicting_values']}, found {len(conflict_values)}",
    )
    require(
        len(conflict_decisions) == inventory.expected_counts["conflict_decisions"],
        f"conflict decisions: expected {inventory.expected_counts['conflict_decisions']}, found {len(conflict_decisions)}",
    )
    for value in conflict_values:
        require(bool(value.evidence_links), f"conflicting value {value.id} has no evidence lineage")

    by_code = {config.stable_vehicle_code: config for config in configs}
    readiness_statuses: Counter[tuple[str, str]] = Counter()
    readiness_by_key = {
        (row.vehicle_configuration_id, row.vehicle_fitment_id, row.readiness_type): row
        for row in readiness_rows
    }
    for config in configs:
        for fitment in [None, *config.fitments]:
            try:
                evaluations = evaluate_readiness(session, config, fitment=fitment)
            except Exception as exc:  # pragma: no cover - the message is part of the QA report
                failures.append(
                    f"readiness raised for {config.stable_vehicle_code}/"
                    f"{fitment.fitment_code if fitment else 'configuration'}: {exc}"
                )
                continue
            for evaluation in evaluations:
                status = evaluation.status.value
                readiness_statuses[(evaluation.readiness_type.value, status)] += 1
                row = readiness_by_key.get(
                    (config.id, fitment.id if fitment else None, evaluation.readiness_type.value)
                )
                require(
                    row is not None,
                    f"missing persisted readiness for {config.stable_vehicle_code}/{evaluation.readiness_type.value}",
                )
                if row is not None:
                    require(
                        row.status == status,
                        f"persisted readiness mismatch for {config.stable_vehicle_code}/{evaluation.readiness_type.value}",
                    )

    representative: dict[str, Any] = {}

    # The following checks are deliberately small, reusable semantic sentinels.
    # They run when the relevant accepted record is a member of the release.
    byd = by_code.get("th-byd-atto3-my24-extended-local")
    if byd is not None:
        clearances = _value_for(values, byd.id, "clearance_value_mm")
        require({value.numeric_value for value in clearances} == {150.0, 175.0}, "BYD clearance values are not 150/175 mm")
        require(
            {value.load_condition.mass_basis for value in clearances if value.load_condition}
            == {"UNLADEN", "OEM_LADEN"},
            "BYD clearance load conditions are not UNLADEN/OEM_LADEN",
        )
        representative["BYD ATTO 3"] = {
            "clearance_mm": sorted(value.numeric_value for value in clearances),
            "clearance_load_basis": sorted(
                value.load_condition.mass_basis for value in clearances if value.load_condition
            ),
        }

    triton = by_code.get("th-mitsubishi-triton-ultra-4wd-at-release-2023")
    if triton is not None:
        turning = _value_for(values, triton.id, "turning_radius_normalized_m")
        require(
            triton.model_year_from is None
            and triton.model_year_to is None
            and triton.identity_time_basis == "EDITION_RELEASE",
            "Mitsubishi Triton does not preserve the non-MY EDITION_RELEASE identity",
        )
        require(
            len(turning) == 1
            and turning[0].numeric_value == 6.2
            and turning[0].semantic_metadata
            == {
                "turning_radius_or_diameter": "RADIUS",
                "turning_reference": "OEM_UNSPECIFIED",
                "turning_axle_scope": "OEM_UNSPECIFIED",
            },
            "Mitsubishi Triton turning semantics are not preserved",
        )
        representative["Mitsubishi Triton"] = {
            "identity_time_basis": triton.identity_time_basis,
            "model_year_from": triton.model_year_from,
            "turning_radius_m": turning[0].numeric_value if turning else None,
            "turning_reference": turning[0].semantic_metadata.get("turning_reference") if turning else None,
        }

    volvo = by_code.get("th-volvo-ex30-ultra-smer-my2026-19")
    if volvo is not None:
        volvo_fitment_codes = {fitment.fitment_code for fitment in volvo.fitments}
        width_codes = {
            "overall_width_body_mm",
            "overall_width_including_mirrors_mm",
            "overall_width_mirrors_folded_mm",
        }
        width_values = [
            value
            for value in values
            if value.vehicle_configuration_id == volvo.id
            and value.parameter_definition
            and value.parameter_definition.parameter_code in width_codes
        ]
        turning_text = _value_for(values, volvo.id, "oem_turning_value_text")
        radius = _value_for(values, volvo.id, "turning_radius_normalized_m")
        volvo_decisions = [
            decision for decision in conflict_decisions if decision.vehicle_configuration_id == volvo.id
        ]
        require(volvo_fitment_codes == {"WHEEL19"}, "Volvo EX30 WHEEL19 fitment scope is missing")
        require(
            {value.parameter_definition.parameter_code for value in width_values} == width_codes
            and all(value.vehicle_fitment_id is None for value in width_values),
            "Volvo EX30 body/mirrors-open/mirrors-folded widths are not distinct configuration values",
        )
        require(
            any(
                value.load_condition is not None
                and value.load_condition.name == "Kerb weight plus one person"
                for value in values
                if value.vehicle_configuration_id == volvo.id
                and value.parameter_definition
                and value.parameter_definition.parameter_code in {"overall_height_mm", "clearance_value_mm"}
            ),
            "Volvo EX30 kerb+1 load scope is missing",
        )
        require(
            len(turning_text) == 2
            and {value.text_value for value in turning_text} == {"10.7 m", "11 m"}
            and all(value.resolution_state == "CONFLICTING" for value in turning_text)
            and not radius
            and not volvo_decisions,
            "Volvo EX30 turning conflict/unknown semantics are not preserved",
        )
        representative["Volvo EX30"] = {
            "fitment_codes": sorted(volvo_fitment_codes),
            "turning_text_values": sorted(value.text_value for value in turning_text),
            "turning_radius_normalized": bool(radius),
            "conflict_decisions": len(volvo_decisions),
        }

    for code in (
        "th-honda-civic-ehev-rs-release-2026-07-23",
        "th-honda-accord-ehev-rs-release-2025-08-22",
    ):
        config = by_code.get(code)
        if config is None:
            continue
        b_values = [
            value
            for value in values
            if value.vehicle_configuration_id == config.id and value.authority_grade == "B"
        ]
        require(bool(b_values), f"{code} has no authority-B values")
        require(
            all(
                {
                    link.source_observation.source_document.authority_class
                    for value in b_values
                    for link in value.evidence_links
                    if link.source_observation and link.source_observation.source_document
                }
                == {"REPUTABLE_SECONDARY"}
                for _ in [0]
            ),
            f"{code} authority-B values are not attributable only to REPUTABLE_SECONDARY evidence",
        )
        representative[code] = {"authority_B_values": len(b_values), "source_authority": "REPUTABLE_SECONDARY"}

    for code in (
        "th-tesla-model3-premium-long-range-rwd-2024plus",
        "th-tesla-modely-premium-lr-rwd-19-2025plus",
    ):
        config = by_code.get(code)
        if config is None:
            continue
        text_values = _value_for(values, config.id, "oem_turning_value_text")
        radius = _value_for(values, config.id, "turning_radius_normalized_m")
        assessment = _assessment_for(assessments, config.id, "turning_radius_normalized_m")
        require(
            bool(text_values) and not radius and bool(assessment),
            f"{code} turning-circle raw text/radius assessment is not preserved",
        )
        representative[code] = {
            "raw_turning_text": [value.text_value for value in text_values],
            "normalized_radius": bool(radius),
        }

    mg = by_code.get("th-mg-im6-long-range-release-2025-08-22")
    if mg is not None:
        mg_observations = [observation for observation in observations if observation.vehicle_configuration_id == mg.id]
        rear_steer_observation = next(
            (observation for observation in mg_observations if "four-wheel steering" in observation.raw_value.lower()),
            None,
        )
        mg_relations = [
            relation
            for relation in session.scalars(select(SteeringRelation)).all()
            if relation.vehicle_configuration_id == mg.id
        ]
        require(rear_steer_observation is not None, "MG IM6 rear/four-wheel steering source observation is missing")
        require(not mg_relations, "MG IM6 has invented rear-steering kinematics")
        representative["MG IM6"] = {
            "four_wheel_steering_observation": rear_steer_observation.raw_value if rear_steer_observation else None,
            "steering_relations": len(mg_relations),
        }

    lexus = by_code.get("th-lexus-lm350h-executive-7seat-allnew-2023")
    if lexus is not None:
        turning = _value_for(values, lexus.id, "turning_radius_normalized_m")
        linked_labels = {
            link.source_observation.raw_label
            for value in turning
            for link in value.evidence_links
            if link.source_observation
        }
        require(
            len(turning) == 1
            and turning[0].numeric_value == 5.9
            and (turning[0].semantic_metadata or {}).get("turning_reference") == "WHEEL_PATH_OTHER"
            and (turning[0].semantic_metadata or {}).get("turning_axle_scope") == "OEM_UNSPECIFIED"
            and "Minimum Turning Radius (Tire)" in linked_labels,
            "Lexus LM turning semantics are not preserved",
        )
        representative["Lexus LM"] = {
            "turning_radius_m": turning[0].numeric_value if turning else None,
            "turning_reference": turning[0].semantic_metadata.get("turning_reference") if turning else None,
        }

    crv = by_code.get("th-honda-crv-ehev-rs4wd-minorchange-release-2025-11-28")
    if crv is not None:
        height_values = _value_for(values, crv.id, "overall_height_mm")
        height_assessments = _assessment_for(assessments, crv.id, "overall_height_mm")
        require(not height_values, "Honda CR-V height must remain UNKNOWN in this release")
        require(
            any(assessment.availability_state == "UNKNOWN" for assessment in height_assessments),
            "Honda CR-V height UNKNOWN assessment is missing",
        )
        representative["Honda CR-V"] = {"height_values": len(height_values), "height_unknown": bool(height_assessments)}

    serena = by_code.get("th-nissan-serena-epower-highway-star-release-2025-03-24")
    if serena is not None:
        serena_observations = [
            observation
            for observation in observations
            if observation.vehicle_configuration_id == serena.id
        ]
        approximate_weight = next(
            (
                observation
                for observation in serena_observations
                if "1797" in observation.raw_value and "weight" in observation.raw_label.lower()
            ),
            None,
        )
        kerb_values = _value_for(values, serena.id, "kerb_mass_kg")
        kerb_assessments = _assessment_for(assessments, serena.id, "kerb_mass_kg")
        require(approximate_weight is not None, "Nissan Serena approximate 1,797 kg raw observation is missing")
        require(not kerb_values, "Nissan Serena approximate weight was incorrectly normalized as kerb mass")
        require(bool(kerb_assessments), "Nissan Serena kerb-mass UNKNOWN assessment is missing")
        representative["Nissan Serena"] = {
            "approximate_weight_raw": approximate_weight.raw_value if approximate_weight else None,
            "kerb_mass_values": len(kerb_values),
        }

    for value in values:
        code = value.parameter_definition.parameter_code if value.parameter_definition else None
        metadata = value.semantic_metadata or {}
        if code == "overall_width_reported_mm":
            require(
                metadata.get("width_envelope_definition") not in {
                    "BODY_EXCLUDING_MIRRORS",
                    "MIRRORS_OPEN",
                    "MIRRORS_FOLDED",
                },
                f"generic OEM width value {value.id} was promoted to a body/mirror envelope",
            )
        if code in {"oem_front_tread_or_track_mm", "oem_rear_tread_or_track_mm"}:
            require(
                metadata.get("track_definition") not in {"AVT_OUTER_FACE", "AVT_OUTER_FACE_TYRE"},
                f"OEM tread value {value.id} was promoted to AVT outer-face track",
            )

    if failures:
        raise BuildError("database QA failed:\n" + "\n".join(f"- {failure}" for failure in failures))

    return {
        "database_counts": db_counts,
        "readiness_status": {
            f"{readiness_type}:{status}": count
            for (readiness_type, status), count in sorted(readiness_statuses.items())
        },
        "representative_proofs": representative,
        "source_reuse": {
            "manifest_entries": inventory.expected_counts["source_entries"],
            "unique_source_codes": inventory.expected_counts["sources"],
        },
    }


def _export_proof(
    session: Any,
    inventory: ManifestInventory,
    csv_path: Path,
    xlsx_path: Path,
    *,
    root: Path = ROOT,
) -> dict[str, Any]:
    rows = export_rows(session, inventory.stable_vehicle_codes)
    csv_data = csv_bytes(rows)
    xlsx_data = xlsx_bytes(rows)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.write_bytes(csv_data)
    xlsx_path.write_bytes(xlsx_data)

    csv_rows = list(csv.DictReader(csv_data.decode("utf-8-sig").splitlines()))
    workbook = load_workbook(BytesIO(xlsx_data), read_only=True)
    try:
        worksheet = workbook["Engineering Data"]
        headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        xlsx_rows = sum(1 for _ in worksheet.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    failures: list[str] = []
    expected_rows = inventory.expected_counts["values"] + inventory.expected_counts["assessments"]
    if len(csv_rows) != expected_rows:
        failures.append(f"CSV export row count: expected {expected_rows}, found {len(csv_rows)}")
    if xlsx_rows != expected_rows:
        failures.append(f"XLSX export row count: expected {expected_rows}, found {xlsx_rows}")
    if headers != EXPORT_COLUMNS:
        failures.append("XLSX headers do not match EXPORT_COLUMNS")
    if {row["stable_vehicle_code"] for row in csv_rows} != set(inventory.stable_vehicle_codes):
        failures.append("CSV export does not contain exactly the release stable vehicle codes")
    if not all(
        row["value_kind"] == "assessment"
        or (row["source_observation_ids"] and row["source_document_codes"])
        for row in csv_rows
    ):
        failures.append("CSV export contains a source-backed row without observation/document provenance")
    if inventory.expected_counts["fitments"] and not any(row["fitment_code"] for row in csv_rows):
        failures.append("CSV export does not preserve fitment scope")
    if inventory.expected_counts["loads"] and not any(row["load_condition_id"] for row in csv_rows):
        failures.append("CSV export does not preserve load-condition scope")
    if inventory.expected_counts["conflicting_values"] and not any(
        row["resolution_state"] == "CONFLICTING" for row in csv_rows
    ):
        failures.append("CSV export does not preserve conflict state")
    if b"password=" in csv_data.lower() or b"password=" in xlsx_data.lower():
        failures.append("export appears to contain a credential query field")
    if failures:
        raise BuildError("export proof failed:\n" + "\n".join(f"- {failure}" for failure in failures))

    return {
        "csv_path": _portable_relative(csv_path, root),
        "xlsx_path": _portable_relative(xlsx_path, root),
        "csv_rows": len(csv_rows),
        "vehicles": len({row["stable_vehicle_code"] for row in csv_rows}),
        "csv_bytes": len(csv_data),
        "xlsx_bytes": len(xlsx_data),
        "headers_match": headers == EXPORT_COLUMNS,
        "provenance_present": True,
        "fitment_scope_present": any(row["fitment_code"] for row in csv_rows),
        "load_scope_present": any(row["load_condition_id"] for row in csv_rows),
        "conflict_state_present": any(row["resolution_state"] == "CONFLICTING" for row in csv_rows),
        "credentials_exposed": False,
    }


def _git_revision(root: Path) -> str:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=root,
            check=True,
            capture_output=True,
            text=True,
        )
    except (OSError, subprocess.CalledProcessError):
        return "UNKNOWN"
    return result.stdout.strip() or "UNKNOWN"


def _write_json(path: Path, document: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(document, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def metadata_path_for(database_path: Path) -> Path:
    """Return the local metadata sidecar path for one database file."""

    database_path = Path(database_path)
    return database_path.with_name(database_path.name + ".meta.json")


def database_sha256(database_path: Path) -> str:
    """Hash one database file without opening or mutating it."""

    try:
        return hashlib.sha256(Path(database_path).read_bytes()).hexdigest()
    except OSError as exc:
        raise BuildError(f"could not hash database {database_path}: {exc}") from exc


def load_database_metadata(database_path: Path) -> dict[str, Any] | None:
    """Load the adjacent metadata sidecar, or None when it is absent."""

    metadata_path = metadata_path_for(database_path)
    if not metadata_path.is_file():
        return None
    value = _load_json_value(metadata_path)
    if not isinstance(value, dict):
        raise BuildError(f"database metadata root must be an object: {metadata_path}")
    return value


def validate_database_metadata(
    database_path: Path,
    *,
    expected_release_id: str | None = None,
    expected_release_definition: str | None = None,
    expected_build_input_digest: str | None = None,
    expected_stable_vehicle_digest: str | None = None,
    expected_vehicle_count: int | None = None,
    require_promoted: bool = False,
) -> dict[str, Any]:
    """Validate that a database and its sidecar describe the same build."""

    database_path = Path(database_path).resolve()
    if not database_path.is_file():
        raise BuildError(f"database file was not found: {database_path}")
    metadata = load_database_metadata(database_path)
    if metadata is None:
        raise BuildError(f"database metadata sidecar was not found: {metadata_path_for(database_path)}")

    required = {
        "metadata_schema_version",
        "release_id",
        "release_definition",
        "build_input_digest_sha256",
        "stable_vehicle_digest_sha256",
        "vehicle_count",
        "database_sha256",
        "promoted_database_sha256",
        "promotion_result",
        "build_utc",
        "promotion_utc",
    }
    missing = sorted(required - set(metadata))
    if missing:
        raise BuildError("database metadata is missing fields: " + ", ".join(missing))
    if metadata["metadata_schema_version"] != PROMOTED_DATABASE_METADATA_SCHEMA_VERSION:
        raise BuildError(
            "unsupported database metadata schema version "
            f"{metadata['metadata_schema_version']!r}; expected "
            f"{PROMOTED_DATABASE_METADATA_SCHEMA_VERSION!r}"
        )
    if require_promoted and metadata["promotion_result"] != "PROMOTED":
        raise BuildError(
            f"database metadata promotion result is {metadata['promotion_result']!r}, expected 'PROMOTED'"
        )
    if metadata["promotion_result"] not in {"PROMOTED", "NOT_REQUESTED"}:
        raise BuildError(
            f"invalid database metadata promotion result: {metadata['promotion_result']!r}"
        )

    actual_sha256 = database_sha256(database_path)
    if metadata["database_sha256"] != actual_sha256:
        raise BuildError(
            "database SHA-256 mismatch against metadata: "
            f"expected {metadata['database_sha256']}, found {actual_sha256}"
        )
    promoted_sha256 = metadata["promoted_database_sha256"]
    if promoted_sha256 != actual_sha256:
        raise BuildError(
            "promoted database SHA-256 mismatch against metadata: "
            f"expected {promoted_sha256}, found {actual_sha256}"
        )

    expected_values = {
        "release_id": expected_release_id,
        "release_definition": expected_release_definition,
        "build_input_digest_sha256": expected_build_input_digest,
        "stable_vehicle_digest_sha256": expected_stable_vehicle_digest,
        "vehicle_count": expected_vehicle_count,
    }
    for field_name, expected in expected_values.items():
        if expected is not None and metadata[field_name] != expected:
            raise BuildError(
                f"database metadata {field_name} mismatch: "
                f"expected {expected!r}, found {metadata[field_name]!r}"
            )
    return metadata


def validate_promoted_database_metadata(
    database_path: Path,
    **expected: Any,
) -> dict[str, Any]:
    """Validate metadata required for an accepted promoted database."""

    return validate_database_metadata(database_path, require_promoted=True, **expected)


def _metadata_document(
    inventory: ManifestInventory,
    *,
    root: Path,
    database_hash: str,
    promotion_result: str,
    timestamp: str | None = None,
) -> dict[str, Any]:
    timestamp = timestamp or datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    return {
        "metadata_schema_version": PROMOTED_DATABASE_METADATA_SCHEMA_VERSION,
        "release_id": inventory.release.release_id,
        "release_definition": _portable_relative(inventory.release.path, root),
        "build_input_digest_sha256": inventory.build_input_digest_sha256,
        "stable_vehicle_digest_sha256": inventory.stable_vehicle_digest,
        "vehicle_count": len(inventory.manifests),
        "database_sha256": database_hash,
        "promoted_database_sha256": database_hash,
        "promotion_result": promotion_result,
        "build_utc": timestamp,
        "promotion_utc": timestamp,
    }


def _promotion_temp_path(path: Path, label: str) -> Path:
    return path.with_name(f".{path.name}.{label}.{os.getpid()}.{uuid.uuid4().hex}.tmp")


def _remove_file(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _accepted_database_snapshot(path: Path, root: Path) -> dict[str, Any] | None:
    """Capture the accepted file before a replacement is attempted."""

    if not path.is_file():
        return None
    snapshot: dict[str, Any] = {
        "path": _portable_relative(path, root),
        "bytes": path.stat().st_size,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "vehicles": None,
    }
    try:
        read_only_uri = f"{path.resolve().as_uri()}?mode=ro"
        with sqlite3.connect(read_only_uri, uri=True) as connection:
            snapshot["vehicles"] = connection.execute(
                "SELECT count(*) FROM vehicle_configuration"
            ).fetchone()[0]
    except (OSError, sqlite3.Error):
        # The file hash remains useful even when an older accepted file cannot
        # be opened by the snapshot reader; replacement still remains atomic.
        pass
    return snapshot


def _qualification_record(
    *,
    inventory: ManifestInventory,
    root: Path,
    staging_path: Path,
    staging_database_sha256: str,
    staging_metadata_path: Path,
    promoted_path: Path | None,
    promoted_metadata_path: Path | None,
    previous_path: Path | None,
    previous_metadata_path: Path | None,
    previous_accepted: dict[str, Any] | None,
    registry_only: dict[str, int],
    qa: dict[str, Any],
    exports: dict[str, Any],
) -> dict[str, Any]:
    return {
        "qualification_schema_version": "1.0",
        "release_id": inventory.release.release_id,
        "release_definition": _portable_relative(inventory.release.path, root),
        "release_date": inventory.release.release_date,
        "release_status": inventory.release.release_status,
        "repository_sha": _git_revision(root),
        "manifest_count": len(inventory.manifests),
        "stable_vehicle_codes": list(inventory.stable_vehicle_codes),
        "stable_vehicle_digest_sha256": inventory.stable_vehicle_digest,
        "build_input_digest_sha256": inventory.build_input_digest_sha256,
        "expected_manifest_counts": inventory.expected_counts,
        "database_counts": qa["database_counts"],
        "registry_only": registry_only,
        "readiness_status": qa["readiness_status"],
        "representative_proofs": qa["representative_proofs"],
        "source_reuse": qa["source_reuse"],
        "exports": exports,
        "staging_database": _portable_relative(staging_path, root),
        "staging_database_sha256": staging_database_sha256,
        "staging_metadata": _portable_relative(staging_metadata_path, root),
        "promotion": {
            "requested": promoted_path is not None,
            "result": "PROMOTED" if promoted_path is not None else "NOT_REQUESTED",
            "database": _portable_relative(promoted_path, root),
            "metadata": _portable_relative(promoted_metadata_path, root),
            "previous_database": _portable_relative(previous_path, root),
            "previous_metadata": _portable_relative(previous_metadata_path, root),
        },
        "previous_accepted_database": previous_accepted,
        "qa_result": "PASS",
        "build_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    }


def _sidecars(path: Path) -> tuple[Path, Path]:
    return (
        path.with_name(path.name + "-wal"),
        path.with_name(path.name + "-shm"),
    )


def _promote_staging(
    staging_path: Path,
    final_path: Path,
    *,
    metadata_document: Mapping[str, Any],
) -> Path | None:
    """Replace the accepted DB and sidecar with rollback-safe file operations."""

    staging_path = Path(staging_path).resolve()
    final_path = Path(final_path).resolve()
    final_path.parent.mkdir(parents=True, exist_ok=True)
    staging_metadata_path = metadata_path_for(staging_path)
    final_metadata_path = metadata_path_for(final_path)
    previous_path = final_path.with_name(final_path.name + ".previous")
    previous_metadata_path = metadata_path_for(previous_path)

    if not staging_path.is_file():
        raise BuildError(f"staging database was not found: {staging_path}")
    staging_hash = database_sha256(staging_path)
    if metadata_document.get("database_sha256") != staging_hash:
        raise BuildError("staging database SHA-256 does not match promotion metadata")
    if metadata_document.get("promoted_database_sha256") not in {None, staging_hash}:
        raise BuildError("staging promoted database SHA-256 does not match promotion metadata")

    live_paths = {
        "staging": staging_path,
        "staging_metadata": staging_metadata_path,
        "final": final_path,
        "final_metadata": final_metadata_path,
        "previous": previous_path,
        "previous_metadata": previous_metadata_path,
    }
    original_files = {name: path.is_file() for name, path in live_paths.items()}
    backups: dict[str, Path] = {}
    pending: list[Path] = []
    try:
        # Back up every file that may be touched before the first live replace.
        # The backups stay beside the targets so copy/replace uses one volume.
        for name, path in live_paths.items():
            if original_files[name]:
                backup = _promotion_temp_path(path, f"backup-{name}")
                shutil.copy2(path, backup)
                backups[name] = backup

        pending_previous = _promotion_temp_path(previous_path, "pending-previous")
        pending.append(pending_previous)
        if final_path.is_file():
            shutil.copy2(final_path, pending_previous)

            old_metadata: dict[str, Any] | None = None
            try:
                old_metadata = load_database_metadata(final_path)
            except BuildError:
                # An invalid old marker must not be copied beside a valid old
                # database.  The accepted pair is still restorable from backup.
                old_metadata = None
            old_hash = database_sha256(final_path)
            pending_previous_metadata = _promotion_temp_path(
                previous_metadata_path,
                "pending-previous-metadata",
            )
            pending.append(pending_previous_metadata)
            if (
                old_metadata is not None
                and old_metadata.get("promotion_result") == "PROMOTED"
                and old_metadata.get("database_sha256") == old_hash
                and old_metadata.get("promoted_database_sha256") == old_hash
            ):
                shutil.copy2(metadata_path_for(final_path), pending_previous_metadata)

        pending_metadata = _promotion_temp_path(final_metadata_path, "pending-metadata")
        pending.append(pending_metadata)
        _write_json(pending_metadata, metadata_document)

        if final_path.is_file():
            os.replace(pending_previous, previous_path)
            if pending_previous_metadata in pending and pending_previous_metadata.is_file():
                os.replace(pending_previous_metadata, previous_metadata_path)
            else:
                _remove_file(previous_metadata_path)
        os.replace(staging_path, final_path)
        os.replace(pending_metadata, final_metadata_path)
        _remove_file(staging_metadata_path)
    except (BuildError, OSError, ValueError) as exc:
        rollback_errors: list[str] = []
        # Restore the complete pre-promotion file set, including the staging
        # input.  This is deliberately conservative: a failed pair update
        # must leave the accepted DB and marker as they were.
        for name, path in live_paths.items():
            backup = backups.get(name)
            try:
                if not original_files[name]:
                    _remove_file(path)
                elif backup is not None:
                    _remove_file(path)
                    shutil.copy2(backup, path)
                else:
                    rollback_errors.append(f"{path}: original backup was not created")
            except (OSError, shutil.Error) as rollback_error:
                rollback_errors.append(f"{path}: {rollback_error}")
        detail = f"database promotion failed: {exc}"
        if rollback_errors:
            detail += "; rollback also failed: " + "; ".join(rollback_errors)
        raise BuildError(detail) from exc
    finally:
        for path in [*pending, *backups.values()]:
            _remove_file(path)

    return previous_path if final_path.is_file() and previous_path.is_file() else None


def _build_unlocked(args: argparse.Namespace, *, root: Path = ROOT) -> dict[str, Any]:
    root = root.resolve()
    release_path = getattr(args, "release", None)
    inventory = collect_inventory(release_path, root=root)
    staging_path = _resolved_path(Path(args.staging), root).resolve()
    final_path = _resolved_path(Path(args.final), root).resolve()
    csv_argument = getattr(args, "csv", None)
    xlsx_argument = getattr(args, "xlsx", None)
    qualification_argument = getattr(args, "qualification", None)
    csv_path = (
        _resolved_path(Path(csv_argument), root).resolve()
        if csv_argument
        else root / f"vehicle_engineering_curated.{inventory.release.release_id}.csv"
    )
    xlsx_path = (
        _resolved_path(Path(xlsx_argument), root).resolve()
        if xlsx_argument
        else root / f"vehicle_engineering_curated.{inventory.release.release_id}.xlsx"
    )
    qualification_path = (
        _resolved_path(Path(qualification_argument), root).resolve()
        if qualification_argument
        else root / "artifacts" / "local" / f"{inventory.release.release_id}.qualification.json"
    )
    staging_metadata_path = metadata_path_for(staging_path)
    final_metadata_path = metadata_path_for(final_path)

    if staging_path == final_path:
        raise BuildError("staging and final database paths must differ")
    previous_accepted = _accepted_database_snapshot(final_path, root)
    if staging_path.exists():
        raise BuildError(
            f"staging database already exists: {staging_path}; remove the failed/disposable staging file before retrying"
        )
    if staging_metadata_path.exists():
        raise BuildError(
            f"staging metadata already exists: {staging_metadata_path}; remove the failed/disposable metadata before retrying"
        )
    if not args.no_promote and final_path.exists() and not args.replace_final:
        raise BuildError(
            f"final database already exists: {final_path}; pass --replace-final only after successful staging QA"
        )
    for sidecar in _sidecars(staging_path):
        if sidecar.exists():
            raise BuildError(f"staging sidecar already exists: {sidecar}")

    staging_path.parent.mkdir(parents=True, exist_ok=True)
    environment = os.environ.copy()
    environment["DATABASE_URL"] = _sqlite_url(staging_path)
    environment["PYTHONUNBUFFERED"] = "1"

    _run_command(["-m", "alembic", "upgrade", "head"], environment, root=root)
    _run_command(["-m", "app.curate", "init"], environment, root=root)
    registry_only = _registry_only_qa(environment["DATABASE_URL"])
    for manifest_path in inventory.paths:
        _run_command(["-m", "app.curate", "validate", str(manifest_path)], environment, root=root)
    for manifest_path in inventory.paths:
        _run_command(["-m", "app.curate", "import", str(manifest_path)], environment, root=root)

    engine = make_engine(_sqlite_url(staging_path))
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    try:
        with factory() as session:
            qa = _qa_database(session, inventory)
            exports = _export_proof(session, inventory, csv_path, xlsx_path, root=root)
    finally:
        engine.dispose()

    for sidecar in _sidecars(staging_path):
        if sidecar.exists():
            raise BuildError(f"refusing to promote while SQLite sidecar exists: {sidecar}")

    staging_database_hash = database_sha256(staging_path)
    build_timestamp = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    staging_metadata = _metadata_document(
        inventory,
        root=root,
        database_hash=staging_database_hash,
        promotion_result="PROMOTED" if not args.no_promote else "NOT_REQUESTED",
        timestamp=build_timestamp,
    )
    _write_json(staging_metadata_path, staging_metadata)

    promoted_to: Path | None = None
    previous_database: Path | None = None
    if not args.no_promote:
        previous_database = _promote_staging(
            staging_path,
            final_path,
            metadata_document=staging_metadata,
        )
        promoted_to = final_path
    previous_metadata = (
        metadata_path_for(previous_database)
        if previous_database is not None and metadata_path_for(previous_database).is_file()
        else None
    )

    record = _qualification_record(
        inventory=inventory,
        root=root,
        staging_path=staging_path,
        staging_database_sha256=staging_database_hash,
        staging_metadata_path=staging_metadata_path,
        promoted_path=promoted_to,
        promoted_metadata_path=final_metadata_path if promoted_to is not None else None,
        previous_path=previous_database,
        previous_metadata_path=previous_metadata,
        previous_accepted=previous_accepted,
        registry_only=registry_only,
        qa=qa,
        exports=exports,
    )
    qualification_error: str | None = None
    try:
        _write_json(qualification_path, record)
    except OSError as exc:
        # The database result is already safely staged/promoted.  Keep this
        # metadata artifact best-effort so an artifact filesystem issue cannot
        # turn a successful atomic database promotion into a reported failure.
        qualification_error = str(exc)
        print(f"WARNING: could not write qualification record {qualification_path}: {exc}", file=sys.stderr)

    return {
        "status": "PASS",
        "release_id": inventory.release.release_id,
        "release_definition": _portable_relative(inventory.release.path, root),
        "manifest_count": len(inventory.manifests),
        "stable_vehicle_codes": list(inventory.stable_vehicle_codes),
        "stable_vehicle_digest_sha256": inventory.stable_vehicle_digest,
        "build_input_digest_sha256": inventory.build_input_digest_sha256,
        "expected_manifest_counts": inventory.expected_counts,
        "staging_database": _portable_relative(staging_path, root),
        "staging_database_sha256": staging_database_hash,
        "staging_metadata": _portable_relative(staging_metadata_path, root),
        "promoted_database": _portable_relative(promoted_to, root),
        "promoted_metadata": _portable_relative(
            final_metadata_path if promoted_to is not None else None,
            root,
        ),
        "previous_database": _portable_relative(previous_database, root),
        "previous_metadata": _portable_relative(previous_metadata, root),
        "registry_only": registry_only,
        "qualification_record": _portable_relative(qualification_path, root),
        "qualification_write_error": qualification_error,
        **qa,
        "exports": exports,
    }


def build(
    args: argparse.Namespace,
    *,
    root: Path = ROOT,
    lock_held: bool = False,
) -> dict[str, Any]:
    """Build and promote while serializing mutations of the accepted DB pair."""

    root = root.resolve()
    if args.no_promote or lock_held:
        return _build_unlocked(args, root=root)
    final_path = _resolved_path(Path(args.final), root).resolve()
    with database_refresh_lock(final_path):
        return _build_unlocked(args, root=root)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build the accepted curated database from a release definition")
    parser.add_argument(
        "--release",
        type=Path,
        default=None,
        help=(
            "versioned accepted release definition; when omitted, resolve "
            "data/curation/releases/current_release.json"
        ),
    )
    parser.add_argument(
        "--staging",
        type=Path,
        default=Path("vehicle_engineering_curated.staging.db"),
        help="clean disposable staging database path",
    )
    parser.add_argument(
        "--final",
        type=Path,
        default=Path("vehicle_engineering_curated.db"),
        help="accepted local database path",
    )
    parser.add_argument("--csv", type=Path, default=None, help="CSV proof output path")
    parser.add_argument("--xlsx", type=Path, default=None, help="XLSX proof output path")
    parser.add_argument(
        "--qualification",
        type=Path,
        default=None,
        help="qualification JSON output path (default: artifacts/local/<release-id>.qualification.json)",
    )
    parser.add_argument(
        "--no-promote",
        action="store_true",
        help="leave the successful staging DB in place for inspection",
    )
    parser.add_argument(
        "--replace-final",
        action="store_true",
        help="allow promotion to replace an existing accepted database after all gates pass",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    try:
        report = build(_parser().parse_args(argv))
    except (BuildError, OSError, ValueError) as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
